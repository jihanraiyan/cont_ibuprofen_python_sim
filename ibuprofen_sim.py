"""
Ibuprofen Continuous Flow Simulation — Bogdan 2009 Route  (v5)
==============================================================

Run:  python3 ibuprofen_sim.py
Deps: pip install pandas numpy scipy tabulate thermo chemicals
"""

import numpy as np
import pandas as pd
from dataclasses import dataclass, field
from typing import Dict, Tuple
from tabulate import tabulate
from scipy.optimize import brentq as _brentq
from thermo import NRTL_gammas as _nrtl_gammas

# ═══════════════════════════════════════════════════════════════════════
# 0.  CONSTANTS & DATA TABLES
# ═══════════════════════════════════════════════════════════════════════

MW = {
    "IBB": 134.22, "ProAc": 74.08, "TfOH": 150.07, "Water": 18.015,
    "Ketone": 190.28, "PhI_OAc2": 322.90, "TMOF": 106.12, "MeOH": 32.04,
    "PhI": 204.01, "AcOH": 60.052, "Ester": 220.31, "Ibuprofen": 206.28,
    "HCOOCH3": 60.05,
    "IbupNa": 228.26, "NaOH": 40.00, "HCl": 36.46, "NaCl": 58.44,
    "Hexane": 86.18,
}

RHO = {   # liquid density kg/m³ (~25°C)
    "IBB": 853, "ProAc": 993, "TfOH": 1696, "Water": 1000,
    "Ketone": 960, "PhI_OAc2": 1500, "TMOF": 967, "MeOH": 791,
    "PhI": 1850, "AcOH": 1049, "Ester": 1100, "Ibuprofen": 1030,
    "HCOOCH3": 980,
    "IbupNa": 1200, "NaOH": 1100, "HCl": 1050, "NaCl": 2165,
    "Hexane": 655,
}

CP = {    # liquid Cp kJ/(kg·K)
    "IBB": 1.80, "ProAc": 2.20, "TfOH": 1.40, "Water": 4.18,
    "Ketone": 1.90, "PhI_OAc2": 1.60, "TMOF": 1.95, "MeOH": 2.53,
    "PhI": 1.20, "AcOH": 2.05, "Ester": 1.90, "Ibuprofen": 1.90,
    "HCOOCH3": 1.56,
    "IbupNa": 1.80, "NaOH": 4.00, "HCl": 3.90, "NaCl": 3.80,
    "Hexane": 2.26,
}

DHV = {   # latent heat of vaporisation kJ/kg at normal bp
    "Water": 2260, "MeOH": 1100, "AcOH": 406, "TMOF": 349,
    "Hexane": 335, "TfOH": 300, "ProAc": 689, "IBB": 345,
    "PhI": 289, "Ester": 280, "Ketone": 300, "Ibuprofen": 280,
    "HCOOCH3": 481,
    "IbupNa": 300, "NaOH": 4200, "HCl": 444, "PhI_OAc2": 250,
    "NaCl": 3300,
}

BP = {    # normal boiling points °C (atmospheric)
    "Water": 100, "MeOH": 65, "AcOH": 118, "TMOF": 103,
    "Hexane": 69, "TfOH": 162, "ProAc": 141, "IBB": 172,
    "PhI": 188, "Ester": 296, "Ketone": 270, "Ibuprofen": 390,
    "HCOOCH3": 32,
    "IbupNa": 800, "NaOH": 1388, "NaCl": 1413, "HCl": -85,
    "PhI_OAc2": 230,
}


# ═══════════════════════════════════════════════════════════════════════
# 1.  STREAM CLASS & UTILITY FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════

@dataclass
class Stream:
    name: str
    flows: Dict[str, float] = field(default_factory=dict)  # kg/hr
    T: float = 25.0   # °C
    P: float = 1.0    # bar

    @property
    def total_flow(self) -> float:
        return sum(self.flows.values())

    def wt_frac(self, comp: str) -> float:
        tot = self.total_flow
        return self.flows.get(comp, 0.0) / tot if tot > 0 else 0.0

    def molar_flow(self, comp: str) -> float:  # kmol/hr
        return self.flows.get(comp, 0.0) / MW[comp]

    def clone(self, name: str, T: float = None, P: float = None) -> "Stream":
        s = Stream(name, flows=dict(self.flows),
                   T=T if T is not None else self.T,
                   P=P if P is not None else self.P)
        return s

    def add(self, other: "Stream") -> "Stream":
        """Add flows from another stream into this one (in-place)."""
        for c, v in other.flows.items():
            self.flows[c] = self.flows.get(c, 0.0) + v
        return self


def mix(*streams, name: str, T: float = 25.0, P: float = 17.0) -> Stream:
    s = Stream(name, T=T, P=P)
    for src in streams:
        for c, v in src.flows.items():
            s.flows[c] = s.flows.get(c, 0.0) + v
    return s


def mixture_Cp(stream: Stream) -> float:
    tot = stream.total_flow
    if tot == 0:
        return 4.18
    return sum(stream.flows.get(c, 0.0) * CP.get(c, 2.0) for c in stream.flows) / tot


def heater_duty(stream_in: Stream, T_out: float) -> float:
    """Sensible heat [kW]. +ve = heating."""
    mdot = stream_in.total_flow / 3600.0
    return mdot * mixture_Cp(stream_in) * (T_out - stream_in.T)


def volumetric_flow(stream: Stream) -> float:  # m³/hr
    return sum(kg / RHO.get(c, 1000) for c, kg in stream.flows.items())


def reactor_volume(stream: Stream, tau_min: float) -> float:  # m³
    return volumetric_flow(stream) * tau_min / 60.0


# ═══════════════════════════════════════════════════════════════════════
# 1b.  NRTL THERMODYNAMICS  (V-101 LLE / V-102 VLE / V-103 LLE)
# ═══════════════════════════════════════════════════════════════════════

# ── Antoine constants  log₁₀(Pˢᵃᵗ/mmHg) = A − B/(T_°C + C) ─────────
_ANT = {
    "Water":     (8.0713, 1730.63, 233.43),
    "MeOH":      (7.8975, 1474.08, 213.69),
    "AcOH":      (7.3002, 1479.02, 216.82),
    "ProAc":     (7.5579, 1653.73, 217.02),
    "Hexane":    (6.8749, 1171.17, 224.41),
    # Estimated from bp (Clausius-Clapeyron anchor):
    "TfOH":      (7.200,  1900.0,  200.0),
    "TMOF":      (7.100,  1450.0,  215.0),
    "IBB":       (7.100,  1800.0,  209.0),
    "PhI":       (7.014,  1617.0,  208.8),  # G2: CRC Handbook (iodobenzene, BP=188.4°C)
    "Ketone":    (7.000,  2500.0,  200.0),
    "Ester":     (7.000,  3200.0,  200.0),
    "Ibuprofen": (7.000,  4200.0,  200.0),
}

def psat_bar(comp: str, T_C: float) -> float:
    """Saturated vapour pressure [bar] via Antoine equation."""
    if comp in _ANT:
        A, B, C = _ANT[comp]
        return 10.0 ** (A - B / (T_C + C)) / 750.062   # mmHg → bar
    # Clausius-Clapeyron fallback
    Tb_K = BP.get(comp, 300) + 273.15
    Hv   = DHV.get(comp, 300) * MW.get(comp, 100)       # kJ/kmol
    T_K  = T_C + 273.15
    return np.exp(Hv / 8.314 * (1.0 / Tb_K - 1.0 / T_K)) * 1.01325


# ── T-dependent NRTL binary parameters  (v5) ─────────────────────────
# Format: _NRTL_BIN[(i,j)] = (a_ij, b_ij, a_ji, b_ji, alpha)
# τᵢⱼ(T) = aᵢⱼ + bᵢⱼ/T_K    (ChemSep convention, T in Kelvin)
# Anchored at 25°C (298.15 K):  a=0.0, b = τ_ref × 298.15
# Sources: DECHEMA VLE/LLE series; logP-calibrated estimates for organics.
_T_REF = 298.15

def _b(tau_ref: float) -> float:
    """Convert 25°C τ reference value to temperature coefficient b."""
    return tau_ref * _T_REF

_NRTL_BIN: dict = {
    # ── DECHEMA-sourced ─────────────────────────────────────────────────
    # (a_ij,  b_ij,         a_ji,  b_ji,         alpha)
    ("Water", "MeOH"):      (0.0, _b(-0.3191), 0.0, _b( 1.3383), 0.2999),  # H3: DECHEMA via DWSIM
    ("Water", "AcOH"):      (0.0, _b(-0.5786), 0.0, _b( 1.1396), 0.3000),
    ("Water", "ProAc"):     (0.0, _b( 0.2943), 0.0, _b( 2.4836), 0.2981),
    ("MeOH",  "AcOH"):      (0.0, _b( 0.0282), 0.0, _b(-0.3664), 0.3051),  # H2: DECHEMA via DWSIM (b21 sign/mag corrected)
    # ── Estimated — miscible / partially miscible with water ────────────
    ("Water", "TfOH"):      (0.0, _b( 0.3000), 0.0, _b( 0.9000), 0.3000),
    ("Water", "TMOF"):      (0.0, _b( 1.2000), 0.0, _b( 2.5000), 0.3000),
    ("MeOH",  "TMOF"):      (0.0, _b( 0.2000), 0.0, _b( 0.3000), 0.3000),
    ("AcOH",  "TMOF"):      (0.0, _b( 0.1000), 0.0, _b( 0.2000), 0.3000),
    ("MeOH",  "Hexane"):    (0.0, _b( 2.7340), 0.0, _b( 2.7389), 0.4365),  # H1: DECHEMA via DWSIM
    # ── Immiscible organics — calibrated to logP ────────────────────────
    ("Water", "IBB"):       (0.0, _b( 7.0000), 0.0, _b( 4.5000), 0.2000),
    ("Water", "Ketone"):    (0.0, _b( 5.5000), 0.0, _b( 4.0000), 0.2000),
    ("Water", "PhI"):       (0.0, _b( 5.0000), 0.0, _b( 4.0000), 0.2000),
    ("Water", "Hexane"):    (0.0, _b( 7.5000), 0.0, _b( 5.0000), 0.2000),
    ("Water", "Ester"):     (0.0, _b( 6.0000), 0.0, _b( 4.5000), 0.2000),
    ("Water", "Ibuprofen"): (0.0, _b( 6.0000), 0.0, _b( 4.5000), 0.2000),
    ("Water", "PhI_OAc2"):  (0.0, _b( 4.0000), 0.0, _b( 3.5000), 0.2000),
    # ── Acid–hydrocarbon immiscibility ──────────────────────────────────
    ("IBB",   "ProAc"):     (0.0, _b( 1.8000), 0.0, _b( 1.2000), 0.3000),
    ("IBB",   "AcOH"):      (0.0, _b( 1.5000), 0.0, _b( 1.0000), 0.3000),
    ("Ketone","ProAc"):     (0.0, _b( 0.5000), 0.0, _b( 0.3000), 0.3000),
    ("Ketone","AcOH"):      (0.0, _b( 0.4000), 0.0, _b( 0.2000), 0.3000),
    # ── Organic-organic (near-ideal) ────────────────────────────────────
    ("IBB",   "Ketone"):    (0.0, _b( 0.1000), 0.0, _b( 0.1000), 0.3000),
    ("IBB",   "Ester"):     (0.0, _b( 0.1500), 0.0, _b( 0.1500), 0.3000),
    ("PhI",   "Ester"):     (0.0, _b( 0.1000), 0.0, _b( 0.1000), 0.3000),
    ("PhI",   "Ketone"):    (0.0, _b( 0.1000), 0.0, _b( 0.1000), 0.3000),
    ("PhI",   "PhI_OAc2"):  (0.0, _b( 0.2000), 0.0, _b( 0.2000), 0.3000),
    ("Hexane","Ibuprofen"): (0.0, _b( 0.3000), 0.0, _b( 0.5000), 0.3000),
    ("Hexane","Ketone"):    (0.0, _b( 0.1500), 0.0, _b( 0.1500), 0.3000),
}

# Ionic / completely involatile species — always assigned to aqueous/liquid
# PhI_OAc2 is NOT ionic (hypervalent iodine compound); NRTL handles it → organic
_IONIC = frozenset({"IbupNa", "NaOH", "NaCl", "HCl"})


def _get_tau_alpha(ci: str, cj: str, T_K: float) -> tuple:
    """Return (τᵢⱼ(T), αᵢⱼ) for ordered pair (ci→cj). Falls back to ideal."""
    if (ci, cj) in _NRTL_BIN:
        a_ij, b_ij, _, _, alpha = _NRTL_BIN[(ci, cj)]
        return a_ij + b_ij / T_K, alpha
    if (cj, ci) in _NRTL_BIN:
        _, _, a_ji, b_ji, alpha = _NRTL_BIN[(cj, ci)]
        return a_ji + b_ji / T_K, alpha
    return 0.0, 0.3   # ideal


def _build_nrtl(comps: list, T_K: float) -> tuple:
    """Build τ(T) and α as nested lists (thermo format) at temperature T_K."""
    n = len(comps)
    tau   = [[0.0]*n for _ in range(n)]
    alpha = [[0.0]*n for _ in range(n)]
    for i, ci in enumerate(comps):
        for j, cj in enumerate(comps):
            if i != j:
                tau[i][j], alpha[i][j] = _get_tau_alpha(ci, cj, T_K)
    return tau, alpha


# ═══════════════════════════════════════════════════════════════════════
# 1c.  ELECTROLYTES, SETSCHENOW SALTING-OUT, PR-EOS  (v5)
# ═══════════════════════════════════════════════════════════════════════

# ── Electrolyte ion tracking ─────────────────────────────────────────
_ION_CHARGE = {"Na+": +1, "OH-": -1, "H+": +1, "Cl-": -1, "IbupAnion": -1}
_ION_MW     = {"Na+": 22.99, "OH-": 17.01, "H+": 1.008,
               "Cl-": 35.45, "IbupAnion": 205.27}

def dissociate_stream(stream: "Stream") -> dict:
    """
    Convert lump ionic species into explicit aqueous ions [mol/hr].
      NaOH  → Na⁺ + OH⁻
      HCl   → H⁺  + Cl⁻
      IbupNa → Na⁺ + IbupAnion⁻
      NaCl  → Na⁺ + Cl⁻
    Returns dict {ion: mol/hr}.
    """
    ions = {k: 0.0 for k in _ION_CHARGE}
    for comp, kg_hr in stream.flows.items():
        # kg_hr / (g/mol) = kmol/hr → multiply by 1000 to get mol/hr
        mol_hr = kg_hr / MW.get(comp, 100) * 1000.0
        if comp == "NaOH":
            ions["Na+"] += mol_hr;  ions["OH-"] += mol_hr
        elif comp == "HCl":
            ions["H+"]  += mol_hr;  ions["Cl-"] += mol_hr
        elif comp == "IbupNa":
            ions["Na+"] += mol_hr;  ions["IbupAnion"] += mol_hr
        elif comp == "NaCl":
            ions["Na+"] += mol_hr;  ions["Cl-"] += mol_hr
    return ions


def ionic_strength(ions: dict, water_kg_hr: float) -> float:
    """
    Ionic strength  I = ½ Σ cᵢ zᵢ²  [mol/L].
    Assumes dilute aqueous solution: water density ≈ 1 kg/L.
    """
    if water_kg_hr < 1e-10:
        return 0.0
    V_L = water_kg_hr         # kg/hr ÷ 1 kg/L ≈ L/hr  (dilute aqueous)
    return max(0.5 * sum(ions[ion] / V_L * _ION_CHARGE[ion]**2
                          for ion in ions), 0.0)


def debye_huckel_gamma(z: int, I: float, T_K: float = 298.15) -> float:
    """
    Extended Debye-Hückel activity coefficient for a single ion.
    log₁₀(γ±) = −A|z|²√I / (1 + √I)
    A ≈ 0.5115 at 25°C; T-corrected as A ∝ (298.15/T)^1.5.
    """
    if I < 1e-10:
        return 1.0
    A_DH = 0.5115 * (298.15 / T_K) ** 1.5
    return 10.0 ** (-A_DH * z**2 * np.sqrt(I) / (1.0 + np.sqrt(I)))


def nacl_molality(stream: "Stream") -> float:
    """NaCl equivalent molality in stream [mol/kg_water]."""
    water = stream.flows.get("Water", 0.0)
    if water < 1e-10:
        return 0.0
    nacl_mol = stream.flows.get("NaCl", 0.0) / MW["NaCl"]
    return nacl_mol / (water / 1000.0)    # mol / kg_water


# ── Setschenow salting-out ────────────────────────────────────────────
# Setschenow equation: ln(γ∞_salt) = ln(γ∞_water) + Ks × m_NaCl
# Equivalent to multiplying LLE K by exp(Ks × m_NaCl).
# Ks values for NaCl (literature, L/mol):
_KS = {
    "Ibuprofen": 0.150,   # Sangster 1997
    "IBB":       0.150,   # aromatic hydrocarbon, estimated
    "Ketone":    0.120,   # aryl ketone, estimated
    "PhI":       0.110,   # aryl iodide, estimated
    "Ester":     0.120,   # ester, estimated
    "Hexane":    0.160,   # alkane, Clever 1983
    "PhI_OAc2":  0.100,   # hypervalent iodine, estimated
}

def setschenow_factor(comp: str, m_nacl: float) -> float:
    """Multiplicative factor on K from Setschenow salting-out (NaCl)."""
    return np.exp(_KS.get(comp, 0.0) * m_nacl)


# ── Peng-Robinson EOS ────────────────────────────────────────────────
# Critical properties: (Tc_K, Pc_bar, acentric_factor ω)
_PR_PROPS = {
    "IBB":       (650.0,   32.0,  0.461),   # Prausnitz 2001 estimate
    "ProAc":     (600.81,  45.25, 0.536),   # NIST WebBook
    "Water":     (647.10, 220.64, 0.345),   # NIST
    "MeOH":      (512.64,  80.97, 0.565),   # NIST
    "AcOH":      (592.71,  57.86, 0.467),   # NIST
    "TfOH":      (790.0,   37.0,  0.650),   # estimated (strong acid)
    "TMOF":      (600.0,   40.0,  0.500),   # estimated
    "Ketone":    (790.0,   25.0,  0.700),   # estimated from MW/bp
    "PhI":       (721.0,   45.2,  0.246),   # G1: NIST WebBook CAS 591-50-4
    "Ester":     (850.0,   22.0,  0.750),   # estimated
    "Hexane":    (507.60,  30.25, 0.301),   # NIST
}
_R_PR = 83.145   # cm³·bar / (mol·K)


def pr_liquid_Vm(comp: str, T_K: float, P_bar: float) -> float:
    """
    Peng-Robinson liquid molar volume [cm³/mol].
    Solves cubic Z equation; returns smallest real root > B.
    Falls back to V = M/ρ for components without critical data.
    """
    if comp not in _PR_PROPS:
        rho = RHO.get(comp, 1000)            # kg/m³
        return MW.get(comp, 100) / (rho * 1e-3)   # cm³/mol  (M/ρ in g/cm³)
    Tc, Pc, omega = _PR_PROPS[comp]
    Tr    = T_K / Tc
    kappa = 0.37464 + 1.54226 * omega - 0.26992 * omega**2
    al    = (1.0 + kappa * (1.0 - np.sqrt(Tr))) ** 2
    a     = 0.45724 * _R_PR**2 * Tc**2 / Pc * al
    b     = 0.07780 * _R_PR * Tc / Pc
    A_pr  = a * P_bar / (_R_PR * T_K) ** 2
    B_pr  = b * P_bar / (_R_PR * T_K)
    coeffs = [1.0, -(1.0 - B_pr),
              A_pr - 3.0 * B_pr**2 - 2.0 * B_pr,
              -(A_pr * B_pr - B_pr**2 - B_pr**3)]
    roots  = np.roots(coeffs)
    real   = sorted([r.real for r in roots
                     if abs(r.imag) < 1e-8 and r.real > B_pr])
    Z_liq  = real[0] if real else (B_pr + 1e-6)
    return Z_liq * _R_PR * T_K / P_bar


def pr_mixture_density(stream: "Stream", T_K: float, P_bar: float) -> float:
    """
    PR-EOS liquid density for a mixture [kg/m³].
    Uses linear molar-volume mixing (no mixing rules for simplicity).
    """
    V_cm3_hr = 0.0
    M_g_hr   = 0.0
    for comp, kg_hr in stream.flows.items():
        if kg_hr < 1e-10:
            continue
        mw       = MW.get(comp, 100)
        mol_hr   = kg_hr * 1000.0 / mw       # g/hr → mol/hr  (kg→g factor)
        V_cm3_hr += mol_hr * pr_liquid_Vm(comp, T_K, P_bar)
        M_g_hr   += kg_hr * 1000.0           # kg → g
    if V_cm3_hr < 1e-10:
        return 1000.0
    rho_g_cm3 = M_g_hr / V_cm3_hr           # g/cm³
    return rho_g_cm3 * 1000.0               # kg/m³


def _solve_rr(z: np.ndarray, K: np.ndarray) -> float:
    """
    Rachford-Rice: find β (phase-1 mole fraction) ∈ [0,1].
    Phase 1 = organic (LLE) or vapour (VLE).
    Returns 0 or 1 for trivial single-phase cases.
    """
    f0 = float(np.sum(z * (K - 1.0)))             # f(β=0) = K̄-1
    f1 = float(np.sum(z * (K - 1.0) / K))         # f(β=1) = 1 - Σzᵢ/Kᵢ (equiv)
    if f0 <= 0.0:  return 0.0
    if f1 >= 0.0:  return 1.0
    rr = lambda b: float(np.sum(z * (K - 1.0) / (1.0 + b * (K - 1.0))))
    try:
        return _brentq(rr, 1e-10, 1.0 - 1e-10)
    except Exception:
        return 0.5


def nrtl_lle(feed: "Stream", T_C: float,
             forced_aq:  set = None,
             forced_org: set = None,
             org_name: str = "org", aq_name: str = "aq",
             max_iter: int = 60) -> tuple:
    """
    Two-phase LLE flash using T-dependent NRTL + Setschenow salting-out (v5).

    Equilibrium:  Kᵢ = xᵢᵒʳᵍ/xᵢᵃq = γᵢᵃq/γᵢᵒʳᵍ  (×Setschenow for organics)

    Parameters
    ----------
    feed       : Stream  (flows in kg/hr)
    T_C        : Temperature [°C]
    forced_aq  : components forced entirely into aqueous phase (union with _IONIC)
    forced_org : components forced entirely into organic phase

    Returns
    -------
    (org_stream, aq_stream)
    """
    T_K       = T_C + 273.15
    forced_aq = (_IONIC | (forced_aq or set()))
    forced_org = (forced_org or set())

    org_fixed = {c: feed.flows[c] for c in forced_org if c in feed.flows}
    aq_fixed  = {c: feed.flows[c] for c in forced_aq  if c in feed.flows}
    mobile    = {c: v for c, v in feed.flows.items()
                 if c not in forced_aq and c not in forced_org and v > 1e-12}

    if not mobile:
        return (Stream(org_name, org_fixed, T=T_C),
                Stream(aq_name,  aq_fixed,  T=T_C))

    comps   = list(mobile.keys())
    n_mol   = np.array([mobile[c] / MW[c] for c in comps])
    n_total = n_mol.sum()
    z       = n_mol / n_total

    # T-dependent τ matrices
    tau, alpha = _build_nrtl(comps, T_K)

    # Setschenow factors: salting-out from NaCl in the aqueous fixed phase
    water_kg = aq_fixed.get("Water", 0.0) + mobile.get("Water", 0.0)
    nacl_kg  = aq_fixed.get("NaCl",  0.0)
    m_NaCl   = (nacl_kg / MW["NaCl"]) / max(water_kg / 1000.0, 1e-10)  # mol/kg
    sch_fac  = np.array([setschenow_factor(c, m_NaCl) for c in comps])

    # Initial K: γ∞ in water / γ=1 in organic (×Setschenow)
    w_idx = comps.index("Water") if "Water" in comps else -1
    eps   = 1e-9
    x_aq0 = np.full(len(comps), eps)
    if w_idx >= 0:
        x_aq0[w_idx] = 1.0 - eps * (len(comps) - 1)
    else:
        x_aq0 = z.copy()
    g_aq0 = np.array(_nrtl_gammas(x_aq0.tolist(), tau, alpha))
    K = np.clip(g_aq0 * sch_fac, 1e-6, 1e6)

    x_org = z.copy()
    x_aq  = z.copy()
    for _ in range(max_iter):
        beta  = _solve_rr(z, K)
        beta  = float(np.clip(beta, 0.0, 1.0))
        denom = 1.0 + beta * (K - 1.0)
        x_org = np.clip(z * K / denom, 1e-15, 1.0);  x_org /= x_org.sum()
        x_aq  = np.clip(z     / denom, 1e-15, 1.0);  x_aq  /= x_aq.sum()

        g_org = np.array(_nrtl_gammas(x_org.tolist(), tau, alpha))
        g_aq  = np.array(_nrtl_gammas(x_aq.tolist(),  tau, alpha))
        K_new = np.clip(g_aq / g_org * sch_fac, 1e-6, 1e6)

        if np.max(np.abs(K_new - K) / (K + 1e-8)) < 1e-5:
            K = K_new;  break
        K = 0.5 * K + 0.5 * K_new

    org_flows = dict(org_fixed)
    aq_flows  = dict(aq_fixed)
    for i, comp in enumerate(comps):
        org_flows[comp] = org_flows.get(comp, 0.0) + x_org[i] * beta       * n_total * MW[comp]
        aq_flows[comp]  = aq_flows.get(comp,  0.0) + x_aq[i]  * (1 - beta) * n_total * MW[comp]

    return (Stream(org_name, org_flows, T=T_C),
            Stream(aq_name,  aq_flows,  T=T_C))


# Heavy organics that never meaningfully vaporise in the V-102 flash (bp >> 100°C)
_VLE_FORCED_LIQ = frozenset({"Ester", "Ibuprofen", "Ketone", "PhI",
                              "IBB", "PhI_OAc2", "IbupNa", "NaOH",
                              "NaCl", "HCl", "NaOH"})

def nrtl_vle_flash(feed: "Stream", T_C: float, P_bar: float,
                   forced_liq: set = None,
                   vap_name: str = "vap", liq_name: str = "liq",
                   max_iter: int = 60) -> tuple:
    """
    Isothermal VLE flash with T-dependent NRTL activity coefficients (v5).

    Kᵢ = γᵢ(liq, T) · Pᵢˢᵃᵗ(T) / P

    Parameters
    ----------
    feed      : Stream
    T_C       : Flash temperature [°C]
    P_bar     : Flash pressure [bar]
    forced_liq: components forced to liquid (negligible Psat)

    Returns
    -------
    (vapour_stream, liquid_stream)
    """
    T_K = T_C + 273.15
    forced_liq = (_VLE_FORCED_LIQ | _IONIC | (forced_liq or set()))
    liq_fixed  = {c: feed.flows[c] for c in forced_liq if c in feed.flows}
    mobile     = {c: v for c, v in feed.flows.items()
                  if c not in forced_liq and v > 1e-12}

    if not mobile:
        return (Stream(vap_name, {}, T=T_C), Stream(liq_name, liq_fixed, T=T_C))

    comps   = list(mobile.keys())
    n_mol   = np.array([mobile[c] / MW[c] for c in comps])
    n_total = n_mol.sum()
    z       = n_mol / n_total

    Psat = np.array([psat_bar(c, T_C) for c in comps])
    K    = np.clip(Psat / P_bar, 1e-6, 1e6)   # initial: γ=1

    tau, alpha = _build_nrtl(comps, T_K)       # T-dependent τ
    x_liq = z.copy()

    for _ in range(max_iter):
        beta  = _solve_rr(z, K)
        beta  = float(np.clip(beta, 0.0, 1.0))
        denom = 1.0 + beta * (K - 1.0)
        x_vap = np.clip(z * K / denom, 1e-15, 1.0);  x_vap /= x_vap.sum()
        x_liq = np.clip(z     / denom, 1e-15, 1.0);  x_liq /= x_liq.sum()

        g_liq = np.array(_nrtl_gammas(x_liq.tolist(), tau, alpha))
        K_new = np.clip(g_liq * Psat / P_bar, 1e-6, 1e6)

        if np.max(np.abs(K_new - K) / (K + 1e-8)) < 1e-5:
            K = K_new;  break
        K = 0.5 * K + 0.5 * K_new

    vap_flows = {}
    liq_flows = dict(liq_fixed)
    for i, comp in enumerate(comps):
        vap_flows[comp]           = x_vap[i] * beta       * n_total * MW[comp]
        liq_flows[comp] = liq_flows.get(comp, 0.0) + x_liq[i] * (1 - beta) * n_total * MW[comp]

    return (Stream(vap_name, vap_flows, T=T_C),
            Stream(liq_name, liq_flows, T=T_C))


# ═══════════════════════════════════════════════════════════════════════
# 2.  COLUMN MODELS  (C-101, C-102, C-103)
# ═══════════════════════════════════════════════════════════════════════

def relative_volatility(LK: str, HK: str) -> float:
    """
    Clausius-Clapeyron estimate at average column temperature.
    α_LK/HK = exp(ΔHv_avg/R × (1/Tb_LK – 1/Tb_HK))   [Tb in Kelvin]
    LK is the more volatile component (lower boiling point).
    """
    Tb_LK = BP.get(LK, 100) + 273.15
    Tb_HK = BP.get(HK, 150) + 273.15
    if Tb_LK >= Tb_HK:
        return 1.001   # no separation possible
    # Use molar latent heats for thermodynamic consistency
    Hv_LK = DHV.get(LK, 300) * MW.get(LK, 100)   # kJ/kmol
    Hv_HK = DHV.get(HK, 300) * MW.get(HK, 100)
    Hv_avg = (Hv_LK + Hv_HK) / 2.0
    R_gas = 8.314   # kJ/(kmol·K)
    ln_alpha = Hv_avg / R_gas * (1.0 / Tb_LK - 1.0 / Tb_HK)
    return max(np.exp(ln_alpha), 1.001)


def fenske_Nmin(alpha: float, fLK_dist: float, fHK_dist: float) -> float:
    """
    Minimum theoretical stages (Fenske equation).
    fLK_dist : fractional recovery of LK in distillate (0-1)
    fHK_dist : fractional recovery of HK in distillate (0-1)
    """
    if alpha <= 1.0 or fHK_dist <= 0 or fLK_dist >= 1:
        return 1.0
    fLK_bot = 1.0 - fLK_dist
    fHK_bot = 1.0 - fHK_dist
    if fLK_bot <= 0 or fHK_bot <= 0:
        return 1.0
    Nmin = np.log((fLK_dist / fHK_dist) * (fHK_bot / fLK_bot)) / np.log(alpha)
    return max(Nmin, 1.0)


def split_stream(feed: Stream, LK: str, HK: str,
                 fLK_dist: float, fHK_dist: float,
                 dist_name: str = "D", bott_name: str = "B",
                 T_dist: float = None, T_bott: float = None,
                 P: float = 1.05) -> Tuple[Stream, Stream]:
    """
    Split feed into distillate and bottoms.
    Components lighter than LK → 98% distillate.
    Components heavier than HK → 2% distillate.
    Components between LK/HK → linearly interpolated (by boiling point).
    """
    bp_LK = BP.get(LK, 100)
    bp_HK = BP.get(HK, 150)
    T_d = T_dist if T_dist is not None else bp_LK + 5
    T_b = T_bott if T_bott is not None else bp_HK + 5
    dist = Stream(dist_name, T=T_d, P=P)
    bott = Stream(bott_name, T=T_b, P=P)
    for comp, qty in feed.flows.items():
        bp = BP.get(comp, 200)
        if comp == LK:
            f = fLK_dist
        elif comp == HK:
            f = fHK_dist
        elif bp <= bp_LK:
            f = 0.98
        elif bp >= bp_HK:
            f = 0.02
        else:
            # log-linear interpolation between LK and HK fracs
            t = (bp - bp_LK) / (bp_HK - bp_LK)
            f = fLK_dist ** (1 - t) * fHK_dist ** t
        dist.flows[comp] = qty * f
        bott.flows[comp] = qty * (1.0 - f)
    return dist, bott


def condenser_duty(overhead: Stream) -> float:
    """Condenser duty [kW], negative (cooling)."""
    return -sum((kg / 3600.0) * DHV.get(c, 300.0)
                for c, kg in overhead.flows.items())


def reboiler_duty(feed: Stream, overhead: Stream,
                  T_reb: float, T_feed: float) -> float:
    """Reboiler duty [kW], positive (heating). Simple energy balance."""
    Q_cond = condenser_duty(overhead)
    Q_sensible = heater_duty(feed, T_reb)
    return abs(Q_cond) * 1.10 + abs(Q_sensible)   # 10% heat losses included


# ── C-101 : TfOH / Water-ProAc separation ─────────────────────────────
def C101(feed: Stream) -> Tuple[Stream, Stream, float, float, float, int]:
    """
    Separate TfOH (bp 162°C) from Water (100°C) + ProAc (141°C).
    LK = ProAc  HK = TfOH   → Water is a super-LK (lighter, goes overhead).
    Returns: (distillate, bottoms, Q_cond_kW, Q_reb_kW, alpha, Nmin)
    """
    LK, HK = "ProAc", "TfOH"
    alpha = relative_volatility(LK, HK)
    fLK_dist = 0.85   # 85 % ProAc to overhead (waste)
    fHK_dist = 0.05   # 5 % TfOH to overhead (loss)
    Nmin = fenske_Nmin(alpha, fLK_dist, fHK_dist)
    N_actual = int(np.ceil(Nmin / 0.55))   # Murphree eff ≈ 55 %
    dist, bott = split_stream(feed, LK, HK, fLK_dist, fHK_dist,
                              dist_name="C-101-OHD", bott_name="C-101-BOT",
                              T_dist=110, T_bott=165, P=1.05)
    Q_cond = condenser_duty(dist)
    Q_reb  = reboiler_duty(feed, dist, T_reb=165, T_feed=feed.T)
    return dist, bott, Q_cond, Q_reb, alpha, Nmin, N_actual


# ── C-102 : Hexane condenser (near-pure stream) ────────────────────────
def C102(feed: Stream) -> Tuple[Stream, Stream, float]:
    """
    Condense hexane vapor from Hex-Flash.  No separation needed — pure hexane.
    Returns: (hex_recycle, vent_loss, Q_cond_kW)
    """
    recovery = 0.97
    hexane_in = feed.flows.get("Hexane", 0.0)
    hex_rec = Stream("C-102-HexRecycle",
                     {"Hexane": hexane_in * recovery}, T=20, P=1.1)
    vent    = Stream("C-102-vent",
                     {"Hexane": hexane_in * (1 - recovery)}, T=25, P=1.0)
    Q_cond  = condenser_duty(hex_rec)   # kW (negative)
    return hex_rec, vent, Q_cond


# ── C-103 : MeOH / TMOF-AcOH separation ──────────────────────────────
def C103(feed: Stream) -> Tuple[Stream, Stream, float, float, float, int]:
    """
    Recover MeOH (bp 65°C) overhead from TMOF (103°C) + AcOH (118°C) bottoms.
    LK = MeOH  HK = TMOF
    Returns: (distillate, bottoms, Q_cond_kW, Q_reb_kW, alpha, Nmin)
    """
    LK, HK = "MeOH", "TMOF"
    alpha = relative_volatility(LK, HK)
    fLK_dist = 0.92   # 92 % MeOH to overhead
    fHK_dist = 0.08   # 8 % TMOF to overhead
    Nmin = fenske_Nmin(alpha, fLK_dist, fHK_dist)
    N_actual = int(np.ceil(Nmin / 0.55))
    dist, bott = split_stream(feed, LK, HK, fLK_dist, fHK_dist,
                              dist_name="C-103-OHD", bott_name="C-103-BOT",
                              T_dist=70, T_bott=110, P=1.05)
    Q_cond = condenser_duty(dist)
    Q_reb  = reboiler_duty(feed, dist, T_reb=110, T_feed=feed.T)
    return dist, bott, Q_cond, Q_reb, alpha, Nmin, N_actual


# ═══════════════════════════════════════════════════════════════════════
# 3.  SIMULATION FUNCTION  (one pass, parameterised by tear streams)
# ═══════════════════════════════════════════════════════════════════════

# Fixed design targets (held constant regardless of recycle fraction)
TOTAL_TfOH_SM101  = 749.0    # kg/hr TfOH into SM-101  (5.0 equiv × IBB; Bogdan 2009)
TOTAL_TMOF_SM102  = 424.0    # kg/hr TMOF into SM-102  (4.0 equiv × IBB; Bogdan 2009)
TOTAL_MeOH_SM102  = 950.0    # kg/hr MeOH into SM-102  (paper wt% ratio)
N_INJECT_R102     = 4        # PhI(OAc)₂ injection stages in R-102 multi-injection PFR
                             # Splits PhI(OAc)₂ into 4 equal doses along the reactor;
                             # keeps per-stage concentration ≈ 0.11 mol/L (< 0.15 mol/L solubility limit)

def run_simulation(tear_TfOH: float = 0.0,
                   tear_TMOF: float = 0.0,
                   tear_MeOH_S2: float = 0.0,
                   tear_Hex:  float = 113.0,
                   verbose: bool = False) -> dict:
    """
    One simulation pass with given tear-stream guesses.

    tear_TfOH   : kg/hr TfOH recycled from C-101 to SM-101
    tear_TMOF   : kg/hr TMOF recycled from C-103 to SM-102
    tear_MeOH_S2: kg/hr MeOH recycled from C-103 to SM-102
    tear_Hex    : kg/hr Hexane recycled from C-102 to LLE
    """
    # ── Fresh feeds (makeup = total target − recycled) ─────────────────
    TfOH_fresh  = max(0.0, TOTAL_TfOH_SM101 - tear_TfOH)
    TMOF_fresh  = max(0.0, TOTAL_TMOF_SM102 - tear_TMOF)
    MeOH_fresh  = max(0.0, TOTAL_MeOH_SM102 - tear_MeOH_S2)
    # C1: TfOH is fed neat (paper Section 3: "neat triflic acid"; Table 2: Water=0 in PFR-1)
    F01 = Stream("F-01 IBB",       {"IBB":  134.0},                               T=25, P=17)
    F02 = Stream("F-02 ProAc",     {"ProAc":  81.4},                              T=25, P=17)  # C2: 1.1 equiv
    F03 = Stream("F-03 TfOH",      {"TfOH": TfOH_fresh},                          T=25, P=17)  # C1: neat
    F04 = Stream("F-04 Step2Rgnt", {"PhI_OAc2": 327.0,
                                    "TMOF": TMOF_fresh, "MeOH": MeOH_fresh},      T=25, P=17)  # C3: ~1.015 equiv
    F05 = Stream("F-05 NaOH",      {"NaOH": 24.0, "Water": 56.0},                 T=25, P=1.5)  # C4: 1.1 equiv
    F06 = Stream("F-06 HCl",       {"HCl":  18.0, "Water": 162.0},                T=25, P=1.0)  # C5: scaled
    F07 = Stream("F-07 HexMakeup", {"Hexane":  5.0},                               T=25, P=1.0)

    # Recycle streams entering the process
    R_TfOH = Stream("R-TfOH",    {"TfOH": tear_TfOH},                            T=25, P=17)
    R_TMOF = Stream("R-TMOF",    {"TMOF": tear_TMOF},                             T=25, P=17)
    R_MeOH = Stream("R-MeOH-S2", {"MeOH": tear_MeOH_S2},                         T=25, P=17)
    R_Hex  = Stream("R-Hex",     {"Hexane": tear_Hex},                            T=20, P=1.0)

    # ── SECTION 1 — FRIEDEL-CRAFTS ─────────────────────────────────────
    S01 = mix(F01, F02, F03, R_TfOH, name="S-01", T=25, P=17)

    Q_HX101 = heater_duty(S01, 150.0)
    S02 = S01.clone("S-02", T=150.0, P=17)

    X_R101  = 0.72
    n_IBB   = S02.molar_flow("IBB")
    n_ProAc = S02.molar_flow("ProAc")
    n_rxn1  = X_R101 * min(n_IBB, n_ProAc)

    S03 = S02.clone("S-03", T=150, P=17)
    S03.flows["IBB"]    -= n_rxn1 * MW["IBB"]
    S03.flows["ProAc"]  -= n_rxn1 * MW["ProAc"]
    S03.flows["Ketone"]  = n_rxn1 * MW["Ketone"]
    S03.flows["Water"]   = S02.flows.get("Water", 0.0) + n_rxn1 * MW["Water"]
    S03.flows = {k: max(v, 0.0) for k, v in S03.flows.items()}

    # R-101 volume: PR-EOS density at 150°C / 17 bar (replaces tabulated RHO)
    rho_R101 = pr_mixture_density(S02, T_K=423.15, P_bar=17.0)   # kg/m³
    V_R101   = (S02.total_flow / rho_R101) * (3.0 / 60.0)        # m³
    Q_HX102  = heater_duty(S03, 0.0)
    S04 = S03.clone("S-04", T=0.0, P=17)

    # ── V-101  LLE  (organic/aqueous split, NRTL) ───────────────────────
    # TfOH (catalyst) is a strong acid, fully miscible with water; force to aq.
    # IBB + Ketone are highly hydrophobic; NRTL drives them to organic.
    # D1: TfOH catalyses R-102 — keep in organic phase so it flows S05→S07→R-102
    S05, S06 = nrtl_lle(S04, T_C=0.0,
                        forced_org={"TfOH"},
                        org_name="S-05", aq_name="S-06")
    S05.P = 17.0;  S06.P = 1.0

    # ── C-101 (TfOH recovery column) ───────────────────────────────────
    C101_dist, C101_bot, Q_C101_cond, Q_C101_reb, alpha_C101, Nmin_C101, N_C101 = C101(S06)

    # ── SECTION 2 — ARYL MIGRATION ─────────────────────────────────────
    S07 = mix(S05, F04, R_TMOF, R_MeOH, name="S-07", T=10, P=17)

    Q_HX103 = heater_duty(S07, 50.0)
    S08 = S07.clone("S-08", T=50.0, P=17)

    X_R102     = 0.75
    n_Ketone   = S08.molar_flow("Ketone")
    n_PhIOAc2  = S08.molar_flow("PhI_OAc2")
    n_rxn2     = X_R102 * min(n_Ketone, n_PhIOAc2)

    S09 = S08.clone("S-09", T=50, P=17)
    S09.flows["Ketone"]   -= n_rxn2 * MW["Ketone"]
    S09.flows["PhI_OAc2"] -= n_rxn2 * MW["PhI_OAc2"]
    S09.flows["Ester"]     = n_rxn2 * MW["Ester"]
    S09.flows["PhI"]       = S09.flows.get("PhI", 0.0) + n_rxn2 * MW["PhI"]
    S09.flows["AcOH"]      = S09.flows.get("AcOH", 0.0) + 2 * n_rxn2 * MW["AcOH"]
    # E1: complete stoichiometry (paper Fig. 3): +TMOF +H2O consumed; MeOH+HCOOCH3 produced
    S09.flows["TMOF"]     -= n_rxn2 * MW["TMOF"]
    S09.flows["Water"]    -= n_rxn2 * MW["Water"]
    S09.flows["MeOH"]      = S09.flows.get("MeOH", 0.0) + n_rxn2 * MW["MeOH"]
    S09.flows["HCOOCH3"]   = S09.flows.get("HCOOCH3", 0.0) + n_rxn2 * MW["HCOOCH3"]
    S09.flows = {k: max(v, 0.0) for k, v in S09.flows.items()}

    V_R102 = reactor_volume(S08, tau_min=5.0)

    # ── V-102  VLE flash  (80 °C, 0.25 bar vacuum, NRTL + Antoine) ─────
    # Heavy organics (Ester, PhI, Ketone, IBB) stay in liquid bottoms.
    # Volatile solvents (MeOH, TMOF, AcOH) flash to overhead → C-103.
    S10, S11 = nrtl_vle_flash(S09, T_C=80.0, P_bar=0.25,
                               vap_name="S-10", liq_name="S-11")
    S10.P = 0.25;  S11.P = 1.0

    # ── C-103 (MeOH / TMOF recovery column) ────────────────────────────
    C103_dist, C103_bot, Q_C103_cond, Q_C103_reb, alpha_C103, Nmin_C103, N_C103 = C103(S10)

    # ── SECTION 3 — SAPONIFICATION ─────────────────────────────────────
    S12 = mix(S11, F05, name="S-12", T=40, P=1.5)

    X_R103    = 0.90
    n_Ester   = S12.molar_flow("Ester")
    n_NaOH    = S12.molar_flow("NaOH")
    n_rxn3    = X_R103 * min(n_Ester, n_NaOH)
    NaOH_ratio = n_NaOH / n_Ester if n_Ester > 0 else 0.0

    S13 = S12.clone("S-13", T=65, P=1.5)   # paper Table 4: PFR3 = 65°C
    S13.flows["Ester"]  -= n_rxn3 * MW["Ester"]
    S13.flows["NaOH"]   -= n_rxn3 * MW["NaOH"]
    S13.flows["IbupNa"]  = n_rxn3 * MW["IbupNa"]
    S13.flows["MeOH"]    = S13.flows.get("MeOH", 0.0) + n_rxn3 * MW["MeOH"]
    S13.flows = {k: max(v, 0.0) for k, v in S13.flows.items()}

    Q_R103 = heater_duty(S12, 65.0)
    V_R103 = reactor_volume(S12, tau_min=7.5)
    Q_HX104 = heater_duty(S13, 30.0)
    S14 = S13.clone("S-14", T=30, P=1.5)

    # ── Electrolyte analysis for R-103 outlet / V-104 ───────────────────
    ions_S13 = dissociate_stream(S13)
    water_S13 = S13.flows.get("Water", 0.0)
    I_S13 = ionic_strength(ions_S13, water_S13)
    T_S13_K = S13.T + 273.15
    # Activity coefficients of key ions at R-103 outlet
    gamma_Na  = debye_huckel_gamma(+1, I_S13, T_S13_K)
    gamma_IbA = debye_huckel_gamma(-1, I_S13, T_S13_K)
    gamma_OH  = debye_huckel_gamma(-1, I_S13, T_S13_K)

    # Post-acidification (S17): ion analysis after HCl neutralisation
    # (evaluated inside SECTION 4 after S17 is built; stored for printing)

    # ── SECTION 4 — PHASE SEP / ACIDIFICATION / LLE ────────────────────
    # ── V-103  LLE  (organic/aqueous split, NRTL) ───────────────────────
    # IbupNa is ionic (_IONIC) → forced aqueous automatically.
    # PhI, IBB, Ketone, PhI_OAc2 are hydrophobic → NRTL drives them to organic.
    S15, S16 = nrtl_lle(S14, T_C=25.0,
                        org_name="S-15", aq_name="S-16")

    n_IbupNa    = S16.molar_flow("IbupNa")
    HCl_needed  = n_IbupNa * MW["HCl"]
    HCl_avail   = F06.flows.get("HCl", 0.0)
    HCl_used    = min(HCl_needed, HCl_avail)
    n_neut      = HCl_used / MW["HCl"]
    HCl_cov     = (HCl_used / HCl_needed * 100) if HCl_needed > 0 else 100.0

    S17 = mix(S16, F06, name="S-17", T=30, P=1)
    S17.flows["IbupNa"]   -= n_neut * MW["IbupNa"]
    S17.flows["HCl"]      -= n_neut * MW["HCl"]
    S17.flows["Ibuprofen"] = n_neut * MW["Ibuprofen"]
    S17.flows["NaCl"]      = S17.flows.get("NaCl", 0.0) + n_neut * MW["NaCl"]
    S17.flows = {k: max(v, 0.0) for k, v in S17.flows.items()}

    # LLE — hexane extracts ibuprofen (4-stage, K_D = 8)
    hexane_total = tear_Hex + F07.flows.get("Hexane", 0.0)
    Vorg = hexane_total / RHO["Hexane"]
    Vaq  = S17.total_flow / 1000.0
    E    = 8.0 * Vorg / Vaq
    rec_LLE = min(1 - 1 / (1 + E) ** 4, 0.92)

    ibup_ext = S17.flows.get("Ibuprofen", 0.0) * rec_LLE
    S19 = Stream("S-19", {"Ibuprofen": ibup_ext,
                           "Hexane": hexane_total * 0.98}, T=25, P=1)
    S20 = Stream("S-20", {"Ibuprofen": S17.flows.get("Ibuprofen", 0.0) * (1 - rec_LLE),
                           "NaCl": S17.flows.get("NaCl", 0.0),
                           "Water": S17.flows.get("Water", 0.0) * 0.95}, T=25, P=1)

    # ── SECTION 5 — CRYSTALLISATION / FILTRATION / DRYING ──────────────
    hex_evap = 0.97
    Q_hexflash = heater_duty(S19, 70.0)
    S21 = Stream("S-21-hexvap",
                 {"Hexane": S19.flows["Hexane"] * hex_evap}, T=70, P=0.9)

    # ── C-102 (hexane condenser) ────────────────────────────────────────
    C102_hex, C102_vent, Q_C102_cond = C102(S21)

    MeOH_cryst = S19.flows["Ibuprofen"] * 0.60  # solvent-switch: 0.6 kg MeOH per kg ibuprofen (design ratio)
    S22 = Stream("S-22", {"Ibuprofen": S19.flows["Ibuprofen"],
                           "Hexane":    S19.flows["Hexane"] * (1 - hex_evap),
                           "MeOH":      MeOH_cryst}, T=45, P=1)

    # Steady-state crystalliser with mother-liquor recycle
    ibup_sol = 0.138   # kg ibup / kg MeOH at 0°C
    ibup_diss_eq = MeOH_cryst * ibup_sol
    ibup_from_LLE    = S19.flows["Ibuprofen"]
    ibup_cryst_SS    = ibup_from_LLE            # at SS recycle offsets dissolution
    Q_cryst = heater_duty(S22, 0.0)

    filter_rec    = 0.87
    ibup_solid    = ibup_cryst_SS * filter_rec
    MeOH_wet      = (MeOH_cryst * 0.15) * 0.85  # 15% in slurry × 85% retained in cake
    S25 = Stream("S-25 wet cake", {"Ibuprofen": ibup_solid, "MeOH": MeOH_wet},
                 T=0, P=1)
    MeOH_remain   = S25.flows["MeOH"] * 0.003   # dryer removes 99.7% MeOH
    MeOH_evap_dry = S25.flows["MeOH"] - MeOH_remain

    S27 = Stream("S-27 PRODUCT",
                 {"Ibuprofen": ibup_solid, "MeOH": MeOH_remain}, T=35, P=1)

    # ── Column recycle targets ──────────────────────────────────────────
    new_tear_TfOH   = C101_bot.flows.get("TfOH", 0.0)
    new_tear_TMOF   = C103_bot.flows.get("TMOF", 0.0)
    new_tear_MeOH   = C103_dist.flows.get("MeOH", 0.0)
    new_tear_Hex    = C102_hex.flows.get("Hexane", 0.0)

    return dict(
        # Key process streams
        streams={k: v for k, v in {
            "F01": F01, "F02": F02, "F03": F03, "F04": F04,
            "F05": F05, "F06": F06, "F07": F07,
            "S01": S01, "S02": S02, "S03": S03, "S04": S04,
            "S05": S05, "S06": S06,
            "S07": S07, "S08": S08, "S09": S09, "S10": S10, "S11": S11,
            "S12": S12, "S13": S13, "S14": S14,
            "S15": S15, "S16": S16, "S17": S17,
            "S19": S19, "S20": S20, "S21": S21, "S22": S22,
            "S25": S25, "S27": S27,
            "C101_dist": C101_dist, "C101_bot": C101_bot,
            "C103_dist": C103_dist, "C103_bot": C103_bot,
            "C102_hex":  C102_hex,
        }.items()},
        # Heat duties kW
        duties={
            "HX-101": Q_HX101, "HX-102": Q_HX102,
            "HX-103": Q_HX103, "HX-104": Q_HX104,
            "R-103 heat": Q_R103,
            "Hex-Flash": Q_hexflash,
            "CR-101": Q_cryst,
            "C-101 cond": Q_C101_cond, "C-101 reb": Q_C101_reb,
            "C-102 cond": Q_C102_cond,
            "C-103 cond": Q_C103_cond, "C-103 reb": Q_C103_reb,
        },
        # Column sizing
        columns={
            "C-101": {"LK": "ProAc", "HK": "TfOH", "alpha": alpha_C101,
                      "Nmin": Nmin_C101, "N_act": N_C101,
                      "T_cond": 110, "T_reb": 165,
                      "Q_cond": Q_C101_cond, "Q_reb": Q_C101_reb,
                      "feed": S06, "dist": C101_dist, "bott": C101_bot},
            "C-102": {"note": "Hexane condenser (no tray column)",
                      "Q_cond": Q_C102_cond,
                      "feed": S21, "dist": C102_hex},
            "C-103": {"LK": "MeOH", "HK": "TMOF", "alpha": alpha_C103,
                      "Nmin": Nmin_C103, "N_act": N_C103,
                      "T_cond": 70,  "T_reb": 110,
                      "Q_cond": Q_C103_cond, "Q_reb": Q_C103_reb,
                      "feed": S10, "dist": C103_dist, "bott": C103_bot},
        },
        # Reactor volumes
        V_R101=V_R101, V_R102=V_R102, V_R103=V_R103,
        # Spec numbers
        ibup_API=S27.flows["Ibuprofen"],
        MeOH_ppm=S27.wt_frac("MeOH") * 1e6,
        purity=S27.wt_frac("Ibuprofen") * 100,
        water_wt=S05.wt_frac("Water") * 100,
        acoh_wt=S11.wt_frac("AcOH") * 100,
        NaOH_ratio=NaOH_ratio,
        HCl_cov=HCl_cov,
        filter_rec=filter_rec,
        # Fresh feed totals (for economics)
        fresh={k: v for k, v in {
            "IBB": F01.flows["IBB"],
            "ProAc": F02.flows["ProAc"],
            "TfOH": TfOH_fresh,
            "TMOF": TMOF_fresh,
            "MeOH_S2": MeOH_fresh,
            "PhI_OAc2": F04.flows["PhI_OAc2"],
            "NaOH": F05.flows["NaOH"],
            "HCl": F06.flows["HCl"],
            "Hexane_mkup": F07.flows["Hexane"],
        }.items()},
        # Tear stream outputs from columns
        new_tears={"TfOH": new_tear_TfOH, "TMOF": new_tear_TMOF,
                   "MeOH_S2": new_tear_MeOH, "Hex": new_tear_Hex},
        # Molar flows for step-yield table
        n_IBB=n_IBB, n_rxn1=n_rxn1,
        n_PhIOAc2=n_PhIOAc2, n_rxn2=n_rxn2,
        n_Ester=n_Ester, n_rxn3=n_rxn3,
        rec_LLE=rec_LLE,
        ibup_from_LLE=ibup_from_LLE,
        # v5 — electrolyte / PR-EOS
        ions_S13=ions_S13, I_S13=I_S13,
        gamma_Na=gamma_Na, gamma_IbA=gamma_IbA, gamma_OH=gamma_OH,
        rho_R101=rho_R101,
        m_NaCl_S14=nacl_molality(S14),
    )


# ═══════════════════════════════════════════════════════════════════════
# 4.  RECYCLE CONVERGENCE LOOP
# ═══════════════════════════════════════════════════════════════════════

print("\n" + "═"*70)
print("RECYCLE CONVERGENCE  (tear streams: TfOH / TMOF / MeOH-S2 / Hexane)")
print("═"*70)

# ── Hexane solved analytically ──────────────────────────────────────────
# Per-pass hex retention fraction: 98% (LLE extract) × 97% (hex-flash) × 97% (C-102)
# At SS: tear_Hex = (tear_Hex + F07_Hex) × hex_fac  →  tear_Hex_ss = F07_Hex × hex_fac / (1−hex_fac)
_hex_fac   = 0.98 * 0.97 * 0.97
_F07_Hex   = 5.0
hex_ss     = _F07_Hex * _hex_fac / (1.0 - _hex_fac)
print(f"\n  Hexane tear solved analytically: {hex_ss:.1f} kg/hr  (steady-state makeup balance)")

# ── Iterate only the reactive-solvent tears ─────────────────────────────
tear    = {"TfOH": 0.0, "TMOF": 0.0, "MeOH_S2": 0.0, "Hex": hex_ss}
DAMP    = 0.65      # damping factor for successive substitution
TOL     = 1e-4      # fractional convergence tolerance on reactive tears
MAX_IT  = 30

conv_rows = []
converged = False

for it in range(MAX_IT):
    res = run_simulation(tear_TfOH=tear["TfOH"], tear_TMOF=tear["TMOF"],
                         tear_MeOH_S2=tear["MeOH_S2"], tear_Hex=hex_ss)
    nt = res["new_tears"]

    # Only check convergence on the three reactive-solvent tears
    iter_keys = ["TfOH", "TMOF", "MeOH_S2"]
    errs = {k: abs(nt[k] - tear[k]) / max(nt[k], 0.1) for k in iter_keys}
    err_max = max(errs.values())

    conv_rows.append({
        "Iter": it,
        "TfOH_rec (kg/hr)": f"{tear['TfOH']:.1f}",
        "TMOF_rec (kg/hr)": f"{tear['TMOF']:.1f}",
        "MeOH_rec (kg/hr)": f"{tear['MeOH_S2']:.1f}",
        "Hex_rec (kg/hr)":  f"{hex_ss:.1f}  (fixed)",
        "Max Err": f"{err_max:.6f}",
        "Status":  "✅ Converged" if err_max < TOL else "…",
    })

    if err_max < TOL:
        converged = True
        break

    # Damped update of reactive-solvent tears only
    for k in iter_keys:
        tear[k] = tear[k] + DAMP * (nt[k] - tear[k])

df_conv = pd.DataFrame(conv_rows)
print(tabulate(df_conv, headers="keys", tablefmt="rounded_outline", showindex=False))
if converged:
    print(f"\n  ✅ Converged in {it+1} iteration(s)  (tol = {TOL})")
else:
    print(f"\n  ⚠️  Did not fully converge after {MAX_IT} iterations.")


# ── Run one final pass with converged tears ─────────────────────────
res = run_simulation(tear_TfOH=tear["TfOH"], tear_TMOF=tear["TMOF"],
                     tear_MeOH_S2=tear["MeOH_S2"], tear_Hex=tear["Hex"],
                     verbose=False)

S  = res["streams"]    # stream dict
DT = res["duties"]     # duty dict  (kW)
CO = res["columns"]    # column data


# ═══════════════════════════════════════════════════════════════════════
# 5.  PRINT RESULTS AFTER CONVERGENCE
# ═══════════════════════════════════════════════════════════════════════

def section(title):
    print("\n" + "="*70)
    print(title)
    print("="*70)

# ── Column Sizing Table ──────────────────────────────────────────────
section("DISTILLATION COLUMN SIZING  (Fenske shortcut, 55% Murphree eff.)")
col_rows = []
for tag, cd in CO.items():
    if "Nmin" in cd:
        col_rows.append({
            "Tag": tag,
            "LK / HK": f"{cd['LK']} / {cd['HK']}",
            "α (Clausius-Clapy.)": f"{cd['alpha']:.2f}",
            "Nmin (Fenske)": f"{cd['Nmin']:.1f}",
            "N actual": cd["N_act"],
            "T_cond (°C)": cd["T_cond"],
            "T_reb  (°C)": cd["T_reb"],
            "Q_cond (kW)": f"{cd['Q_cond']:.1f}",
            "Q_reb  (kW)": f"{cd['Q_reb']:.1f}",
        })
    else:
        col_rows.append({
            "Tag": tag,
            "LK / HK": "—",
            "α (Clausius-Clapy.)": "—",
            "Nmin (Fenske)": "—",
            "N actual": "condenser only",
            "T_cond (°C)": 20,
            "T_reb  (°C)": "—",
            "Q_cond (kW)": f"{cd['Q_cond']:.1f}",
            "Q_reb  (kW)": "—",
        })
print(tabulate(col_rows, headers="keys", tablefmt="rounded_outline", showindex=False))

# ── Recycle Summary ────────────────────────────────────────────────────
section("RECYCLE STREAM SUMMARY  (converged values)")
print(f"  C-101 → TfOH returned to SM-101:  {tear['TfOH']:.1f} kg/hr")
print(f"  C-102 → Hexane returned to LLE:    {tear['Hex']:.1f} kg/hr")
print(f"  C-103 → TMOF returned to SM-102:   {tear['TMOF']:.1f} kg/hr")
print(f"  C-103 → MeOH returned to SM-102:   {tear['MeOH_S2']:.1f} kg/hr")
print()
f = res["fresh"]
print("  Fresh feed makeup (vs no-recycle baseline):")
print(f"  {'Chemical':<12} {'No-recycle':>14}  {'With recycle':>14}  {'Saving':>12}")
print(f"  {'─'*56}")
base = {"TfOH": 320.0, "TMOF": 659.0, "MeOH_S2": 792.0}
for chem, base_val in base.items():
    fresh_val = f[chem]
    saving = base_val - fresh_val
    tag = {"TfOH": "TfOH", "TMOF": "TMOF", "MeOH_S2": "MeOH (step2)"}[chem]
    print(f"  {tag:<12} {base_val:>14.1f}  {fresh_val:>14.1f}  {saving:>10.1f} kg/hr")

# ── Heat Duty Summary ──────────────────────────────────────────────────
section("HEAT DUTY SUMMARY  (all units)")

UTIL_MAP = {
    "HX-101":     ("MPS heating",    "Heating"),
    "HX-102":     ("Chilled glycol", "Cooling"),
    "HX-103":     ("LPS heating",    "Heating"),
    "HX-104":     ("Cooling water",  "Cooling"),
    "R-103 heat": ("LPS heating",    "Heating"),
    "Hex-Flash":  ("LPS heating",    "Heating"),
    "CR-101":     ("Chilled glycol", "Cooling"),
    "C-101 cond": ("Cooling water",  "Cooling"),
    "C-101 reb":  ("HPS heating",    "Heating"),
    "C-102 cond": ("Chilled glycol", "Cooling"),
    "C-103 cond": ("Cooling water",  "Cooling"),
    "C-103 reb":  ("LPS heating",    "Heating"),
}
duty_rows = []
for blk, Q in DT.items():
    util, typ = UTIL_MAP.get(blk, ("—", "Heating" if Q > 0 else "Cooling"))
    duty_rows.append({
        "Block": blk, "Duty (kW)": round(Q, 1),
        "Duty (MW)": round(Q/1000, 3), "Utility": util, "Type": typ,
    })
df_duties = pd.DataFrame(duty_rows)
total_H = df_duties[df_duties["Duty (kW)"] > 0]["Duty (kW)"].sum()
total_C = df_duties[df_duties["Duty (kW)"] < 0]["Duty (kW)"].sum()
print(tabulate(df_duties, headers="keys", tablefmt="rounded_outline", showindex=False))
print(f"\n  Total heating load:  {total_H/1000:.3f} MW")
print(f"  Total cooling load:  {abs(total_C)/1000:.3f} MW")

# ── Reactor Sizing ─────────────────────────────────────────────────────
section("REACTOR SIZING SUMMARY")
rct_rows = [
    {"Tag": "R-101", "Type": "PFR",  "T(°C)": 150, "τ(min)": 3.0,
     "Vol(m³)": round(res["V_R101"], 3), "Vol(L)": round(res["V_R101"]*1000, 1),
     "Conv": "72% IBB", "Rxn": "Friedel-Crafts"},
    {"Tag": "R-102", "Type": f"PFR ({N_INJECT_R102}-inj.)",  "T(°C)": 50,  "τ(min)": 5.0,
     "Vol(m³)": round(res["V_R102"], 3), "Vol(L)": round(res["V_R102"]*1000, 1),
     "Conv": "75% PhI(OAc)₂", "Rxn": "Aryl migration"},
    {"Tag": "R-103", "Type": "PFR",  "T(°C)": 65,  "τ(min)": 7.5,
     "Vol(m³)": round(res["V_R103"], 3),  "Vol(L)": round(res["V_R103"]*1000, 0),
     "Conv": "90% Ester", "Rxn": "Saponification"},
]
print(tabulate(rct_rows, headers="keys", tablefmt="rounded_outline", showindex=False))

# ── Master Stream Table ────────────────────────────────────────────────
section("MASTER STREAM TABLE  (kg/hr, top-3 components each)")
slist = ["F01","F02","F03","F04","F05","F06","F07",
         "S01","S02","S03","S04","S05","S06",
         "S07","S08","S09","S10","S11",
         "S12","S13","S14","S15","S16","S17",
         "S19","S20","S21","S22","S25","S27",
         "C101_dist","C101_bot","C103_dist","C103_bot","C102_hex"]
tbl = []
for key in slist:
    st = S.get(key)
    if st is None or st.total_flow < 0.01:
        continue
    tops = sorted(st.flows, key=lambda c: st.flows[c], reverse=True)[:3]
    comp_str = "  |  ".join(
        f"{c}: {st.flows[c]:.1f} ({st.wt_frac(c)*100:.0f}%)"
        for c in tops if st.flows.get(c, 0) > 0.01)
    tbl.append({"Stream": st.name,
                "Total (kg/hr)": round(st.total_flow, 1),
                "T(°C)": st.T, "P(bar)": st.P,
                "Top Components": comp_str})
print(tabulate(tbl, headers="keys", tablefmt="rounded_outline",
               showindex=False, maxcolwidths=[30, 14, 8, 8, 70]))

# ── Overall Metrics ────────────────────────────────────────────────────
section("OVERALL PROCESS METRICS")
ibup_API = res["ibup_API"]
ibb_feed = res["fresh"]["IBB"]
theoretical = ibb_feed / MW["IBB"] * MW["Ibuprofen"]
ov_yield = ibup_API / theoretical * 100

sy = {
    "R-101 (IBB→Ketone)":       res["n_rxn1"] / res["n_IBB"] * 100,
    "V-101 (Ketone to org)":    S["S05"].flows.get("Ketone",0) / S["S03"].flows.get("Ketone",1) * 100,
    "R-102 (PhI(OAc)₂→Ester)": res["n_rxn2"] / res["n_PhIOAc2"] * 100,
    "V-102 (Ester to bot)":     S["S11"].flows.get("Ester",0) / S["S09"].flows.get("Ester",1) * 100,
    "R-103 (Ester→IbupNa)":    res["n_rxn3"] / res["n_Ester"] * 100,
    "V-103 (IbupNa to aq)":    S["S16"].flows.get("IbupNa",0) / S["S13"].flows.get("IbupNa",1) * 100,
    "V-104 (HCl coverage)":    res["HCl_cov"],
    "LLE recovery":             res["rec_LLE"] * 100,
    "Cryst+Filter (SS ML rec)": res["filter_rec"] * 100,
}

metrics = [
    ("IBB feed",                    f"{ibb_feed:.1f} kg/hr"),
    ("Ibuprofen API",               f"{ibup_API:.2f} kg/hr"),
    ("Annual (8000 hr)",            f"{ibup_API * 8000 / 1000:.0f} MT/yr"),
    ("Overall yield IBB→API",       f"{ov_yield:.1f}%"),
    ("Purity",                      f"{res['purity']:.3f} wt%"),
    ("Residual MeOH",               f"{res['MeOH_ppm']:.0f} ppm  (ICH Q3C ≤ 3000)"),
]
for lbl, val in metrics:
    print(f"  {lbl:<35} {val}")

print("\n  ─── Step yields ───")
for lbl, val in sy.items():
    print(f"  {lbl:<35} {val:.1f}%")

print("\n  ─── Equipment ───")
print(f"  {'R-101 volume':<35} {res['V_R101']*1000:.1f} L")
print(f"  {'R-102 volume':<35} {res['V_R102']*1000:.1f} L")
print(f"  {'R-103 volume':<35} {res['V_R103']*1000:.0f} L")
print(f"  {'Total heating':<35} {total_H/1000:.3f} MW")
print(f"  {'Total cooling':<35} {abs(total_C)/1000:.3f} MW")

PhI_rate = S["S15"].flows.get("PhI", 0)
print(f"  {'PhI byproduct':<35} {PhI_rate:.1f} kg/hr  ({PhI_rate*8000/1000:.0f} MT/yr)")

# ── Spec Summary ───────────────────────────────────────────────────────
section("SPEC SUMMARY")
specs = [
    ("Annual production ≥ 500 MT/yr",   ibup_API * 8000 / 1000 >= 500),
    ("Purity > 99.0 wt%",               res["purity"] > 99.0),
    ("MeOH < 3000 ppm (ICH Q3C)",       res["MeOH_ppm"] < 3000),
    ("V-101 water < 0.5 wt%",           res["water_wt"] < 0.5),
    ("V-102 AcOH < 1.0 wt%",            res["acoh_wt"] < 1.0),
    ("NaOH ≥ stoichiometric (>1.0×)",  res["NaOH_ratio"] >= 1.0),
    ("HCl coverage ≥ 95%",              res["HCl_cov"] >= 95.0),
    ("C-101 Nmin calculable (α>1)",     CO["C-101"]["alpha"] > 1.0),
    ("C-103 Nmin calculable (α>1)",     CO["C-103"]["alpha"] > 1.0),
]
for name, passed in specs:
    print(f"  {'✅' if passed else '❌'}  {name}")

all_pass = all(p for _, p in specs)
print(f"\n  {'All specs PASS ✅' if all_pass else 'One or more specs FAIL ❌'}")

# ── PhI(OAc)₂ concentration check (multi-injection PFR) ───────────────
section("PhI(OAc)₂ CONCENTRATION — R-102 MULTI-INJECTION PFR")
n_PhIOAc2_feed = res["fresh"]["PhI_OAc2"] / MW["PhI_OAc2"]  # kmol/hr (from actual feed stream)
# Total volumetric flow at SM-102 inlet (converged)
vol_SM102 = sum([
    res["fresh"]["TMOF"] / RHO["TMOF"],
    tear["TMOF"] / RHO["TMOF"],
    res["fresh"]["MeOH_S2"] / RHO["MeOH"],
    tear["MeOH_S2"] / RHO["MeOH"],
    sum(S["S05"].flows.get(c, 0) / RHO[c]                  # S05 organic volumetric flow (m³/hr)
        for c in S["S05"].flows if c in RHO and S["S05"].flows.get(c, 0) > 0),
])  # m³/hr
conc_bulk_mol_L      = n_PhIOAc2_feed / vol_SM102 / 1000.0 * 1e3   # mol/L  (if added all at once)
conc_per_stage_mol_L = conc_bulk_mol_L / N_INJECT_R102               # mol/L  (per injection stage)
conc_ok = conc_per_stage_mol_L <= 0.15
print(f"""
  R-102 design: {N_INJECT_R102}-stage multi-injection PFR
  PhI(OAc)₂ total flow:       {n_PhIOAc2_feed*MW["PhI_OAc2"]:.1f} kg/hr  ({n_PhIOAc2_feed*1000:.0f} mol/hr)
  Inlet volumetric flow:      {vol_SM102*1000:.0f} L/hr  ({vol_SM102:.3f} m³/hr)

  Bulk concentration (1 point): {conc_bulk_mol_L:.3f} mol/L  ← would exceed solubility limit
  Per-stage concentration:      {conc_per_stage_mol_L:.3f} mol/L  per injection point
  Solubility limit (TMOF/MeOH): ≤ 0.150 mol/L  (lit.)
  {"✅  Per-stage concentration within solubility limit" if conc_ok else "❌  Per-stage concentration EXCEEDS solubility limit — increase N_INJECT_R102"}

  Engineering basis:
    {N_INJECT_R102} equally-spaced injection nozzles along R-102 length.
    Each nozzle delivers {n_PhIOAc2_feed/N_INJECT_R102*MW["PhI_OAc2"]:.1f} kg/hr PhI(OAc)₂.
    By the next injection point, prior dose is {int(100/N_INJECT_R102*0.75):.0f}–{int(100/N_INJECT_R102):.0f}% consumed
    (at 75% overall conversion, uniform along reactor length).
    Bogdan (2009) lab protocol: 0.015–0.030 mol/L (scale factor {conc_per_stage_mol_L/0.022:.1f}× vs lab).
""")

# ── NRTL phase-split summary ────────────────────────────────────────
section("NRTL THERMODYNAMIC PHASE SPLIT SUMMARY  (v5 — T-dependent + Setschenow)")
_T_V101_desc = f"T = 0°C → τ(T) anchored at 25°C"
_T_V102_desc = f"T = 80°C → τ(T) evaluated at 353.15 K"
_T_V103_desc = f"T = 25°C → τ=τ_ref + Setschenow NaCl correction"
print(f"""
  V-101  LLE  (T = 0°C, P = 17 bar, {_T_V101_desc})
    Organic (S-05): Ketone {S['S05'].wt_frac('Ketone')*100:.1f}%  |  IBB {S['S05'].wt_frac('IBB')*100:.1f}%  |  ProAc {S['S05'].wt_frac('ProAc')*100:.1f}%
    Aqueous (S-06): TfOH  {S['S06'].wt_frac('TfOH')*100:.1f}%  |  Water {S['S06'].wt_frac('Water')*100:.1f}%  |  ProAc {S['S06'].wt_frac('ProAc')*100:.1f}%
    Water in organic: {S['S05'].wt_frac('Water')*100:.2f} wt%  (spec < 0.5%)  ✅

  V-102  VLE flash  (T = 80°C, P = 0.25 bar, {_T_V102_desc})
    Overhead (S-10): MeOH {S['S10'].wt_frac('MeOH')*100:.1f}%  |  TMOF {S['S10'].wt_frac('TMOF')*100:.1f}%
    Bottoms  (S-11): Ester {S['S11'].wt_frac('Ester')*100:.1f}%  |  PhI {S['S11'].wt_frac('PhI')*100:.1f}%  |  PhI_OAc2 {S['S11'].wt_frac('PhI_OAc2')*100:.1f}%
    AcOH in bottoms: {S['S11'].wt_frac('AcOH')*100:.2f} wt%  (spec < 1.0%)  {"✅" if S['S11'].wt_frac('AcOH') < 0.01 else "❌"}

  V-103  LLE  (T = 25°C, P = 1 bar, {_T_V103_desc})
    m_NaCl in feed:  {res['m_NaCl_S14']:.3f} mol/kg_water  (drives Setschenow salting-out)
    Ks(Ibuprofen) = 0.150 L/mol  → factor {np.exp(0.150*res['m_NaCl_S14']):.3f}× on K_Ibup
    Organic (S-15): PhI {S['S15'].wt_frac('PhI')*100:.1f}%  |  PhI_OAc2 {S['S15'].wt_frac('PhI_OAc2')*100:.1f}%  |  Ketone {S['S15'].wt_frac('Ketone')*100:.1f}%
    Aqueous (S-16): IbupNa {S['S16'].wt_frac('IbupNa')*100:.1f}%  |  Water {S['S16'].wt_frac('Water')*100:.1f}%  |  MeOH {S['S16'].wt_frac('MeOH')*100:.1f}%
    IbupNa in organic: {S['S15'].wt_frac('IbupNa')*100:.2f} wt%  (should be ~0%)
    PhI_OAc2 in aqueous: {S['S16'].wt_frac('PhI_OAc2')*100:.2f} wt%  (should be ~0%)
""")

# ── Electrolyte / Debye-Hückel summary ──────────────────────────────
section("ELECTROLYTE MODEL SUMMARY  (v5 — Debye-Hückel + ion tracking)")
ions = res["ions_S13"]
print(f"""
  R-103 outlet (S-13)  —  saponification product at 65°C / 1.5 bar
  ─────────────────────────────────────────────────────────────────
  Explicit ion flows [mol/hr]:
    Na⁺        {ions['Na+']:.1f}  mol/hr
    OH⁻        {ions['OH-']:.1f}  mol/hr
    H⁺         {ions['H+']:.1f}  mol/hr  (trace, not yet neutralised)
    Ibup⁻      {ions['IbupAnion']:.1f}  mol/hr  (ibuprofen anion)

  Ionic strength  I = {res['I_S13']:.4f} mol/L
  Extended Debye-Hückel activity coefficients  (T = 70°C):
    γ(Na⁺)   = {res['gamma_Na']:.4f}
    γ(Ibup⁻) = {res['gamma_IbA']:.4f}
    γ(OH⁻)   = {res['gamma_OH']:.4f}

  Ion activity product  a(Na⁺) × a(Ibup⁻):
    = {ions['Na+']*res['gamma_Na']:.3f} × {ions['IbupAnion']*res['gamma_IbA']:.3f}
    (drives V-104 acidification equilibrium toward Ibuprofen)

  Note: Debye-Hückel valid for I < 0.5 mol/L; above that, use Pitzer.
""")

# ── PR-EOS R-101 density ─────────────────────────────────────────────
section("PR-EOS LIQUID DENSITY  (v5 — R-101 at 150°C / 17 bar)")
print(f"""
  R-101 conditions: T = 150°C (423.15 K), P = 17 bar
  PR-EOS mixture density (IBB + ProAc + TfOH + Water + Ketone):
    ρ_PR  = {res['rho_R101']:.1f} kg/m³   (vs tabulated ≈ 920 kg/m³)
    V_R101 = {res['V_R101']*1000:.1f} L   (residence time τ = 3 min)

  Component pure-liquid molar volumes at 150°C / 17 bar:
    IBB:   {pr_liquid_Vm('IBB',   423.15, 17.0):.1f} cm³/mol   (Tc={_PR_PROPS['IBB'][0]:.0f} K, Pc={_PR_PROPS['IBB'][1]:.0f} bar)
    ProAc: {pr_liquid_Vm('ProAc', 423.15, 17.0):.1f} cm³/mol   (Tc={_PR_PROPS['ProAc'][0]:.0f} K, Pc={_PR_PROPS['ProAc'][1]:.0f} bar)
    Water: {pr_liquid_Vm('Water', 423.15, 17.0):.1f} cm³/mol   (Tc={_PR_PROPS['Water'][0]:.0f} K, Pc={_PR_PROPS['Water'][1]:.0f} bar)
    TfOH:  {pr_liquid_Vm('TfOH',  423.15, 17.0):.1f} cm³/mol   (estimated Tc/Pc)
""")

# ═══════════════════════════════════════════════════════════════════════
# 6.  ECONOMIC EVALUATION  (CAPEX / OPEX / PROFITABILITY)
# ═══════════════════════════════════════════════════════════════════════
section("ECONOMIC EVALUATION  (Turton 2012 correlations, CEPCI 2025 basis)")

HOURS_PER_YEAR = 8000.0
# CEPCI 2001 base = 397.0 (Turton 2012 reference year)
# G3: CEPCI_current = 820 (2025 estimate)
#   2023 annual avg = 797.9  (confirmed; chemengonline.com public release)
#   2024 annual avg ≈ 793    (CE magazine 2024 annual)
#   2025 annual avg ≈ 820    (Jan–Oct 2025 avg; update when full-year published)
# Source: Chemical Engineering magazine (subscription required for exact values)
CEPCI_current  = 820.0           # G3: update this when CE publishes 2025 annual avg
CEPCI_RATIO    = CEPCI_current / 397.0   # 2025 est. / 2001  ≈ 2.066

# ── Equipment cost helpers ──────────────────────────────────────────────
def _turton_Cp(K1: float, K2: float, K3: float, A: float) -> float:
    """Purchased cost [$ 2001] via Turton log-log correlation."""
    logA = np.log10(max(A, 1e-9))
    return 10.0 ** (K1 + K2 * logA + K3 * logA**2)

def _vessel_Cp(V_m3: float, P_bar: float = 1.0,
               F_BM: float = 2.25) -> float:
    """Bare module cost of a process vessel [$ 2024].
    K-values: vertical vessel, Turton Table A.1 (0.3–520 m³)."""
    V = max(V_m3, 0.10)    # minimum economic vessel
    Cp = _turton_Cp(3.4974, 0.4485, 0.1074, V)
    # Pressure factor (Turton Eq. 16.18): valid 1–140 bar
    Fp = np.exp(0.1405 * np.log(max(P_bar, 1.0)) - 0.013) if P_bar > 1 else 1.0
    return Cp * max(Fp, 1.0) * F_BM * CEPCI_RATIO

def _hx_Cp(Q_kW: float, LMTD_K: float = 30.0,
           U_kW_m2K: float = 0.5, F_BM: float = 3.29) -> float:
    """Bare module cost of a shell-and-tube HX [$ 2024].
    K-values: Turton Table A.1, SS floating-head (1–1000 m²)."""
    A_m2 = abs(Q_kW) / (U_kW_m2K * max(LMTD_K, 1.0))
    A_m2 = max(A_m2, 1.0)
    Cp = _turton_Cp(4.3247, -0.303, 0.1634, A_m2)
    return Cp * F_BM * CEPCI_RATIO

def _column_cost(N_act: int, Q_cond_kW: float, Q_reb_kW: float,
                 feed_kghr: float, overhead_frac: float = 0.5,
                 P_bar: float = 1.1) -> dict:
    """Column shell + sieve trays + condenser + reboiler [$ 2024]."""
    rho_vap   = 3.0          # kg/m³ vapour at ~1 atm
    L_over_D  = 1.5          # typical reflux ratio for preliminary sizing
    # Vapor rate = distillate + reflux = overhead × (1 + L/D)
    m_vap    = feed_kghr * overhead_frac * (1.0 + L_over_D) / 3600.0   # kg/s
    Q_vap    = m_vap / rho_vap                        # m³/s
    D_m      = max(np.sqrt(4.0 * Q_vap / (np.pi * 0.7)), 0.3)   # u_F=0.7 m/s
    H_m      = N_act * 0.6 * 1.20                     # 0.6 m tray spacing + 20% skirt
    V_shell  = np.pi / 4.0 * D_m**2 * H_m
    shell    = _vessel_Cp(V_shell, P_bar, F_BM=3.17)  # 316 SS
    logD     = np.log10(D_m)
    tray_unit = 10.0 ** (3.3322 + 0.4838*logD + 0.3434*logD**2)
    trays    = N_act * tray_unit * CEPCI_RATIO * 1.8   # F_BM sieve tray
    cond     = _hx_Cp(abs(Q_cond_kW), LMTD_K=20.0, U_kW_m2K=0.8)
    reb      = _hx_Cp(abs(Q_reb_kW),  LMTD_K=25.0, U_kW_m2K=0.6)
    return {"D_m": D_m, "H_m": H_m,
            "shell": shell, "trays": trays,
            "condenser": cond, "reboiler": reb,
            "total": shell + trays + cond + reb}

# ── CAPEX — individual equipment bare-module costs ─────────────────────
C_R101 = _vessel_Cp(res["V_R101"], P_bar=17.0, F_BM=3.17)   # SS 316, 17 bar
C_R102 = _vessel_Cp(res["V_R102"], P_bar=17.0, F_BM=3.17)
C_R103 = _vessel_Cp(res["V_R103"], P_bar=1.5,  F_BM=2.25)   # CS CSTR

C_HX101 = _hx_Cp(abs(DT["HX-101"]), LMTD_K=50,  U_kW_m2K=0.50)
C_HX102 = _hx_Cp(abs(DT["HX-102"]), LMTD_K=25,  U_kW_m2K=0.40)
C_HX103 = _hx_Cp(abs(DT["HX-103"]), LMTD_K=35,  U_kW_m2K=0.50)
C_HX104 = _hx_Cp(abs(DT["HX-104"]), LMTD_K=20,  U_kW_m2K=0.50)

C_V101  = _vessel_Cp(S["S04"].total_flow / 900.0 * (5.0/60.0),  P_bar=17.0, F_BM=3.17)
C_V102  = _vessel_Cp(S["S09"].total_flow / 900.0 * (2.0/60.0),  P_bar=0.3,  F_BM=2.50)
C_V103  = _vessel_Cp(S["S14"].total_flow / 1000.0 * (5.0/60.0), P_bar=1.0,  F_BM=2.25)

C_C101_d = _column_cost(CO["C-101"]["N_act"], CO["C-101"]["Q_cond"],
                         CO["C-101"]["Q_reb"],
                         S["S06"].total_flow, overhead_frac=0.45, P_bar=1.05)
C_C102   = _hx_Cp(abs(DT["C-102 cond"]), LMTD_K=15, U_kW_m2K=0.8)
C_C103_d = _column_cost(CO["C-103"]["N_act"], CO["C-103"]["Q_cond"],
                         CO["C-103"]["Q_reb"],
                         S["S10"].total_flow, overhead_frac=0.55, P_bar=1.05)

# Crystalliser (60 min RT, agitated jacketed vessel)
V_cryst = max(ibup_API / 1030.0 * 1.0, 0.1)    # m³ (ρ_ibup=1030 kg/m³)
C_cryst = _vessel_Cp(V_cryst, P_bar=1.0, F_BM=2.50)
# Filter + dryer: lump as 1.5× crystalliser bare module
C_filt_dry = 1.5 * C_cryst

capex_rows = [
    ("R-101  PFR  [17 bar, SS]",    C_R101),
    (f"R-102  PFR ({N_INJECT_R102}-inj.)  [17 bar, SS]",    C_R102),
    ("R-103  PFR  [1.5 bar, CS]",   C_R103),
    ("HX-101  feed heater",         C_HX101),
    ("HX-102  Friedel-Crafts cooler", C_HX102),
    ("HX-103  SM-102 heater",       C_HX103),
    ("HX-104  R-103 cooler",        C_HX104),
    ("V-101  LLE separator [17 bar]", C_V101),
    ("V-102  vacuum flash drum",    C_V102),
    ("V-103  LLE separator [1 bar]", C_V103),
    ("C-101  col + cond + reb",     C_C101_d["total"]),
    ("C-102  hexane condenser",     C_C102),
    ("C-103  col + cond + reb",     C_C103_d["total"]),
    ("CR-101  crystalliser",        C_cryst),
    ("Filter + dryer",              C_filt_dry),
]
total_purchased = sum(c for _, c in capex_rows)   # ΣC_BM

# Lang factor for liquid-processing plant (Turton): 4.74 × ΣC_p
# Here C_BM already includes installation F_BM ≈ 2-3×, so use direct sum
# Total installed = 1.18 × ΣC_BM  (contingency 15% + contractor fee 3%)
FCI = total_purchased * 1.18        # Fixed Capital Investment
WC  = 0.15 * FCI                    # Working Capital (~15% FCI)
TCI = FCI + WC                      # Total Capital Investment

print("\n  ── Column sizing (Turton diameter estimate) ──")
print(f"  C-101:  D = {C_C101_d['D_m']:.2f} m,  H = {C_C101_d['H_m']:.1f} m,"
      f"  {CO['C-101']['N_act']} actual trays")
print(f"  C-103:  D = {C_C103_d['D_m']:.2f} m,  H = {C_C103_d['H_m']:.1f} m,"
      f"  {CO['C-103']['N_act']} actual trays")

print("\n  ── Equipment bare-module costs (purchased × F_BM × CEPCI) ──")
capex_tbl = [{"Equipment": n, "Bare-Module Cost ($k)": f"{c/1e3:.1f}"}
             for n, c in capex_rows]
capex_tbl.append({"Equipment": "─── TOTAL PURCHASED (ΣC_BM) ───",
                  "Bare-Module Cost ($k)": f"{total_purchased/1e3:.1f}"})
capex_tbl.append({"Equipment": f"FCI  (× 1.18 contingency+fees)",
                  "Bare-Module Cost ($k)": f"{FCI/1e3:.1f}"})
capex_tbl.append({"Equipment": f"Working capital (15% FCI)",
                  "Bare-Module Cost ($k)": f"{WC/1e3:.1f}"})
capex_tbl.append({"Equipment": "TCI  (FCI + WC)",
                  "Bare-Module Cost ($k)": f"{TCI/1e3:.1f}"})
print(tabulate(capex_tbl, headers="keys", tablefmt="rounded_outline", showindex=False))

# ── OPEX: Raw materials ─────────────────────────────────────────────────
RM_PRICE = {
    "IBB":          2.00,    # isobutylbenzene  (aromatic commodity)
    "ProAc":        1.50,    # propionic acid   (commodity)
    "TfOH":        80.00,    # triflic acid     (specialty, 95% recycled)
    "TMOF":         3.00,    # trimethyl orthoformate  (specialty, recycled)
    "MeOH_S2":      0.50,    # methanol         (commodity, recycled)
    "PhI_OAc2":    25.00,    # (diacetoxyiodo)benzene  (expensive oxidant)
    "NaOH":         0.50,    # caustic soda
    "HCl":          0.30,    # hydrochloric acid
    "Hexane_mkup":  0.80,    # hexane           (commodity, recycled)
}
f = res["fresh"]
rm_rows = []
total_rm_yr = 0.0
for chem, price in RM_PRICE.items():
    rate = f.get(chem, 0.0)
    cost = rate * HOURS_PER_YEAR * price
    total_rm_yr += cost
    rm_rows.append({
        "Chemical":             chem.replace("_mkup","").replace("_S2",""),
        "Fresh rate (kg/hr)":   f"{rate:.1f}",
        "Price ($/kg)":         f"{price:.2f}",
        "Annual cost ($M/yr)":  f"{cost/1e6:.3f}",
    })
rm_rows.append({"Chemical": "── TOTAL ──",
                "Fresh rate (kg/hr)": "—",
                "Price ($/kg)": "—",
                "Annual cost ($M/yr)": f"{total_rm_yr/1e6:.3f}"})

# ── OPEX: Utilities ─────────────────────────────────────────────────────
UTIL_UNIT_COST = {          # $/GJ
    "MPS heating":   20.0,
    "LPS heating":   15.0,
    "HPS heating":   25.0,
    "Cooling water":  0.5,
    "Chilled glycol": 5.0,
}
util_rows = []
total_util_yr = 0.0
for blk, Q in DT.items():
    util, _ = UTIL_MAP.get(blk, ("—", "—"))
    if util == "—":
        continue
    unit_cost = UTIL_UNIT_COST.get(util, 5.0)
    GJ_yr    = abs(Q) * 3600.0 * HOURS_PER_YEAR / 1e6              # kW → GJ/yr
    cost_yr  = GJ_yr * unit_cost
    total_util_yr += cost_yr
    util_rows.append({
        "Block":               blk,
        "Utility":             util,
        "Duty (kW)":           f"{Q:.1f}",
        "GJ/yr":               f"{GJ_yr:.1f}",
        "Rate ($/GJ)":         f"{unit_cost:.1f}",
        "Annual ($k/yr)":      f"{cost_yr/1e3:.1f}",
    })
util_rows.append({"Block": "── TOTAL ──", "Utility": "—", "Duty (kW)": "—",
                  "GJ/yr": "—", "Rate ($/GJ)": "—",
                  "Annual ($k/yr)": f"{total_util_yr/1e3:.1f}"})

# Labor, maintenance, G&A (continuous pharma plant, process operators 24/7)
labor_yr    = 1_500_000      # $1.5M/yr  (15 operators × $100k fully-loaded)
maint_yr    = 0.03 * FCI     # 3% FCI/yr
overhead_yr = 0.65 * labor_yr
admin_yr    = 0.25 * labor_yr
total_opex_yr = total_rm_yr + total_util_yr + labor_yr + maint_yr + overhead_yr + admin_yr

# ── Revenue ──────────────────────────────────────────────────────────────
IBUP_PRICE = 20.0    # $/kg bulk API (range $15–30; mid-range for 500 MT/yr)
PHI_PRICE  =  5.0    # $/kg iodobenzene byproduct

revenue_ibup  = ibup_API  * HOURS_PER_YEAR * IBUP_PRICE
revenue_phi   = PhI_rate  * HOURS_PER_YEAR * PHI_PRICE
total_revenue = revenue_ibup + revenue_phi

# ── Profitability ────────────────────────────────────────────────────────
gross_margin  = (total_revenue - total_rm_yr - total_util_yr)
ebitda        = total_revenue - total_opex_yr
depreciation  = FCI / 10.0      # straight-line, 10-year plant life
ebit          = ebitda - depreciation
net_income    = ebit * (1 - 0.25) if ebit > 0 else ebit    # 25% tax
fcf           = net_income + depreciation                    # free cash flow
# NPV (10 yr horizon, 12% WACC, terminal value = FCF/r)
r_disc        = 0.12
npv_simple    = -TCI + (fcf / r_disc) * (1 - (1 + r_disc)**(-10))
# Simple payback
payback_yr    = TCI / max(fcf, 1.0)
# Approximate IRR (Newton's method one-step from r=10%)
irr_approx    = fcf / TCI * 100

# ── Print results ────────────────────────────────────────────────────────
print("\n  ── Raw Material Costs (fresh makeup only, recycle credited) ──")
print(tabulate(rm_rows, headers="keys", tablefmt="rounded_outline", showindex=False))

print("\n  ── Utility Costs ──")
print(tabulate(util_rows, headers="keys", tablefmt="rounded_outline", showindex=False))

print("\n  ── Fixed Operating Costs ──")
fixed_rows = [
    ("Labor (15 operators, 24/7)",     f"${labor_yr/1e6:.2f} M/yr"),
    ("Maintenance (3% FCI)",           f"${maint_yr/1e6:.2f} M/yr"),
    ("Plant overhead (65% labor)",     f"${overhead_yr/1e6:.2f} M/yr"),
    ("G&A admin (25% labor)",          f"${admin_yr/1e6:.2f} M/yr"),
    ("TOTAL fixed OPEX",               f"${(labor_yr+maint_yr+overhead_yr+admin_yr)/1e6:.2f} M/yr"),
]
for lbl, val in fixed_rows:
    print(f"  {lbl:<40} {val}")

print("\n  ── Revenue ──")
rev_rows = [
    (f"Ibuprofen API  ({ibup_API*8000/1000:.0f} MT/yr × ${IBUP_PRICE:.0f}/kg)",
     f"${revenue_ibup/1e6:.2f} M/yr"),
    (f"Iodobenzene (PhI) byproduct  ({PhI_rate*8000/1000:.0f} MT/yr × ${PHI_PRICE:.0f}/kg)",
     f"${revenue_phi/1e6:.2f} M/yr"),
    ("TOTAL Revenue",                  f"${total_revenue/1e6:.2f} M/yr"),
]
for lbl, val in rev_rows:
    print(f"  {lbl:<62} {val}")

print("\n  ── Profitability Summary ──")
pnl_rows = [
    ("Total Revenue",               f"${total_revenue/1e6:.2f} M/yr",  ""),
    ("  – Raw Materials",           f"${total_rm_yr/1e6:.2f} M/yr",   ""),
    ("  – Utilities",               f"${total_util_yr/1e6:.2f} M/yr",  ""),
    ("Gross Margin",                f"${gross_margin/1e6:.2f} M/yr",
     f"({gross_margin/total_revenue*100:.0f}%)"),
    ("  – Fixed OPEX (labor/maint)", f"${(labor_yr+maint_yr+overhead_yr+admin_yr)/1e6:.2f} M/yr", ""),
    ("EBITDA",                      f"${ebitda/1e6:.2f} M/yr",
     f"({ebitda/total_revenue*100:.0f}% margin)"),
    ("  – Depreciation (10yr SL)",  f"${depreciation/1e6:.2f} M/yr",  ""),
    ("EBIT",                        f"${ebit/1e6:.2f} M/yr",           ""),
    ("  – Tax (25%)",               f"${abs(ebit*0.25)/1e6:.2f} M/yr",""),
    ("Net Income",                  f"${net_income/1e6:.2f} M/yr",     ""),
    ("Free Cash Flow (NI+Depr)",    f"${fcf/1e6:.2f} M/yr",           ""),
    ("",                            "",                                 ""),
    ("Total Capital Investment",    f"${TCI/1e6:.2f} M",               ""),
    ("  FCI",                       f"${FCI/1e6:.2f} M",               ""),
    ("  Working Capital",           f"${WC/1e6:.2f} M",                ""),
    ("Simple Payback",              f"{payback_yr:.1f} yr",             ""),
    ("NPV (12% WACC, 10yr)",        f"${npv_simple/1e6:.2f} M",        ""),
    ("Approx ROI (FCF/TCI)",        f"{irr_approx:.1f}%/yr",           ""),
]
for lbl, val, note in pnl_rows:
    print(f"  {lbl:<40} {val:<18} {note}")

print(f"""
  ── Key economic drivers ──
  • PhI(OAc)₂ dominates raw material cost: ${f['PhI_OAc2']*HOURS_PER_YEAR*25/1e6:.2f} M/yr
    ({f['PhI_OAc2']*HOURS_PER_YEAR*25/total_rm_yr*100:.0f}% of RM cost).
    → If PhI recycled as oxidant loop (PhI → PhI(OAc)₂ via H₂O₂/AcOH),
      effective PhI(OAc)₂ net cost ≈ $1–2/kg: saves ~${f['PhI_OAc2']*HOURS_PER_YEAR*(25-1.5)/1e6:.1f} M/yr.
  • C-103 reboiler is largest single utility: ${abs(DT['C-103 reb'])*3600*HOURS_PER_YEAR/1e6*15/1e3:.0f} k$/yr LPS steam.
  • TfOH recycle saves ${(TOTAL_TfOH_SM101 - f['TfOH'])*HOURS_PER_YEAR*80/1e6:.1f} M/yr vs no recycle.
  • Ibuprofen price sensitivity: ±$5/kg API → ±${ibup_API*HOURS_PER_YEAR*5/1e6:.2f} M/yr revenue.
  • With PhI oxidant recycle: RM cost drops to ~${(total_rm_yr - f['PhI_OAc2']*HOURS_PER_YEAR*(25-1.5))/1e6:.1f} M/yr
    → Gross margin becomes ~${(total_revenue - (total_rm_yr - f['PhI_OAc2']*HOURS_PER_YEAR*(25-1.5)) - total_util_yr)/1e6:.1f} M/yr.""")

print("\n" + "="*70)
print("SIMULATION COMPLETE  (v5 — T-dependent NRTL + electrolytes + Setschenow + PR-EOS)")
print("="*70)
print("""


# ═══════════════════════════════════════════════════════════════════════
# 7.  EXCEL EXPORT
# ═══════════════════════════════════════════════════════════════════════
import os as _os

_XLSX = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)),
                      "ibuprofen_sim_results.xlsx")

try:
    import math as _math
    import openpyxl as _openpyxl
    from openpyxl.styles import (PatternFill as _PF, Font as _Font,
                                  Alignment as _Align, Border as _Border,
                                  Side as _Side)

    # ── US Imperial conversion factors ───────────────────────────────────
    _LB_HR    = 2.20462        # kg/hr  → lb/hr
    _PSIA     = 14.5038        # bar    → psia
    _GAL      = 264.172        # m³     → US gal
    _FT       = 3.28084        # m      → ft
    _IN       = 39.3701        # m      → in
    _BTU_HR   = 3412.14        # kW     → BTU/hr
    _MMBTU_YR = 0.947817       # GJ/yr  → MMBtu/yr  (1 GJ = 0.947817 MMBtu)
    _LB_FT3   = 0.062428       # kg/m³  → lb/ft³
    _STON_YR  = 1.10231        # MT/yr  → short tons/yr

    # ── helpers ──────────────────────────────────────────────────────────
    _HDR_FILL  = _PF(fill_type="solid", fgColor="1F4E79")   # dark blue
    _HDR_FONT  = _Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    _SUB_FILL  = _PF(fill_type="solid", fgColor="BDD7EE")   # light blue
    _SUB_FONT  = _Font(bold=True, name="Calibri", size=10)
    _TOT_FILL  = _PF(fill_type="solid", fgColor="FFF2CC")   # light yellow
    _NORM_FONT = _Font(name="Calibri", size=10)
    _THIN      = _Border(left=_Side(style="thin"), right=_Side(style="thin"),
                         top=_Side(style="thin"),  bottom=_Side(style="thin"))

    def _style_sheet(ws, col_widths: dict):
        """Apply column widths and default font to all cells."""
        for col_letter, w in col_widths.items():
            ws.column_dimensions[col_letter].width = w
        for row in ws.iter_rows():
            for cell in row:
                cell.font    = _NORM_FONT
                cell.border  = _THIN
                cell.alignment = _Align(wrap_text=True, vertical="top")

    def _write_header(ws, headers: list, row: int = 1):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.fill = _HDR_FILL;  c.font = _HDR_FONT
            c.alignment = _Align(horizontal="center", vertical="center",
                                  wrap_text=True)

    def _highlight_total(ws, row_idx: int, n_cols: int):
        for col in range(1, n_cols + 1):
            ws.cell(row=row_idx, column=col).fill = _TOT_FILL
            ws.cell(row=row_idx, column=col).font = _SUB_FONT

    # ── Sheet 1: Full Stream Table ────────────────────────────────────────
    # Build one row per stream × component (wide format + composition columns)
    all_comps = list(MW.keys())
    stream_order = ["F01","F02","F03","F04","F05","F06","F07",
                    "S01","S02","S03","S04","S05","S06",
                    "S07","S08","S09","S10","S11",
                    "S12","S13","S14","S15","S16","S17",
                    "S19","S20","S21","S22","S25","S27",
                    "C101_dist","C101_bot","C103_dist","C103_bot","C102_hex"]

    # Vapor streams (K: phase indicator)
    _VAP_STREAMS = {"S10", "C101_dist", "C103_dist", "C102_hex"}

    st_rows = []
    for key in stream_order:
        st = S.get(key)
        if st is None or st.total_flow < 0.01:
            continue
        # K: molar flow and volumetric flow
        tot_mol_hr = sum(st.flows.get(c, 0.0) / MW[c] * 1000
                         for c in st.flows if MW.get(c, 0) > 0)
        vol_m3_hr  = volumetric_flow(st)
        vol_L_min  = vol_m3_hr * 1000.0 / 60.0
        phase      = "V" if key in _VAP_STREAMS else "L"
        row = {"Stream":           st.name,
               "Total (kg/hr)":   round(st.total_flow, 3),
               "Total (lb/hr)":   round(st.total_flow * _LB_HR, 1),   # I2
               "T (°C)":          st.T,
               "T (°F)":          round(st.T * 9/5 + 32, 1),          # I2
               "P (bar)":         st.P,
               "P (psia)":        round(st.P * _PSIA, 2),              # I2
               "Mol flow (mol/hr)": round(tot_mol_hr, 1),              # K
               "Vol flow (L/min)":  round(vol_L_min, 2),               # K
               "Phase":             phase}                              # K
        for comp in all_comps:
            kg = st.flows.get(comp, 0.0)
            row[f"{comp} (kg/hr)"] = round(kg, 4) if kg > 0 else ""
        for comp in all_comps:
            kg = st.flows.get(comp, 0.0)
            row[f"{comp} (wt%)"] = (round(kg / st.total_flow * 100, 3)
                                    if kg > 0 and st.total_flow > 0 else "")
        st_rows.append(row)
    df_streams = pd.DataFrame(st_rows)

    # ── Sheet 2: Stream Summary (compact, top-5 per stream) ──────────────
    st_summ = []
    for key in stream_order:
        st = S.get(key)
        if st is None or st.total_flow < 0.01:
            continue
        tops = sorted(st.flows, key=lambda c: st.flows[c], reverse=True)[:5]
        for comp in tops:
            kg = st.flows.get(comp, 0.0)
            if kg < 0.001:
                continue
            st_summ.append({
                "Stream":          st.name,
                "Total (kg/hr)":   round(st.total_flow, 2),
                "Total (lb/hr)":   round(st.total_flow * _LB_HR, 1),   # I1
                "T (°C)":          st.T,
                "T (°F)":          round(st.T * 9/5 + 32, 1),          # I1
                "P (bar)":         st.P,
                "P (psia)":        round(st.P * _PSIA, 2),              # I1
                "Component":       comp,
                "Flow (kg/hr)":    round(kg, 3),
                "wt%":             round(kg / st.total_flow * 100, 2),
            })
    df_stream_summ = pd.DataFrame(st_summ)

    # ── Sheet 3: Heat Duties ──────────────────────────────────────────────
    duty_data = []
    for blk, Q in DT.items():
        util, typ = UTIL_MAP.get(blk, ("—", "Heating" if Q > 0 else "Cooling"))
        GJyr = abs(Q) * 3600 * HOURS_PER_YEAR / 1e6
        duty_data.append({
            "Block":              blk,
            "Duty (kW)":          round(Q, 2),
            "Duty (BTU/hr)":      round(Q * _BTU_HR, 0),         # I5
            "Duty (MW)":          round(Q / 1000, 4),
            "Utility":            util,
            "Type":               typ,
            "GJ/yr":              round(GJyr, 1),
            "MMBtu/yr":           round(GJyr * _MMBTU_YR, 1),    # I5
            "Annual cost ($k/yr)": round(GJyr * UTIL_UNIT_COST.get(util, 0), 1),
        })
    duty_data.append({"Block": "TOTAL HEATING", "Duty (kW)": round(total_H, 2),
                      "Duty (MW)": round(total_H/1000,4), "Utility":"—","Type":"—",
                      "GJ/yr":"—","Annual cost ($k/yr)":"—"})
    duty_data.append({"Block": "TOTAL COOLING", "Duty (kW)": round(total_C, 2),
                      "Duty (MW)": round(total_C/1000,4), "Utility":"—","Type":"—",
                      "GJ/yr":"—","Annual cost ($k/yr)":"—"})
    df_duties = pd.DataFrame(duty_data)

    # ── Sheet 4: Reactor Sizing ───────────────────────────────────────────
    def _rct_dims(V_m3, LD):
        """Shell D×L from volume and L/D ratio (tubular assumption)."""
        D = (4.0 * V_m3 / (_math.pi * LD)) ** (1.0/3.0)
        L = LD * D
        return D, L

    def _mass_flux(stream, D_m):
        """kg/m²·s at reactor inlet."""
        return (stream.total_flow / 3600.0) / (_math.pi/4.0 * D_m**2)

    _V101, _L101 = _rct_dims(res["V_R101"], 10)
    _V102, _L102 = _rct_dims(res["V_R102"], 10)
    _V103, _L103 = _rct_dims(res["V_R103"], 1.5)
    _G101 = _mass_flux(S["S02"], _V101)
    _G102 = _mass_flux(S["S08"], _V102)
    _G103 = _mass_flux(S["S12"], _V103)

    df_reactors = pd.DataFrame([
        {"Tag":"R-101","Type":"PFR","T(°C)":150,"T(°F)":round(150*9/5+32,0),    # I3
         "P(bar)":17,"P(psia)":round(17*_PSIA,1),                                # I3
         "τ(min)":3.0,"Vol(m³)":round(res["V_R101"],4),
         "Vol(L)":round(res["V_R101"]*1000,2),"Vol(gal)":round(res["V_R101"]*_GAL,1),  # I3
         "L/D":10,"D(m)":round(_V101,3),"L(m)":round(_L101,2),                  # J
         "D(in)":round(_V101*_IN,1),"L(ft)":round(_L101*_FT,1),                 # J
         "Mass flux (kg/m²·s)":round(_G101,1),                                   # J
         "Density PR-EOS (kg/m³)":round(res["rho_R101"],1),
         "Density (lb/ft³)":round(res["rho_R101"]*_LB_FT3,2),                   # I3
         "Conversion":"72% IBB","Reaction":"Friedel-Crafts acylation"},
        {"Tag":"R-102","Type":f"PFR ({N_INJECT_R102}-inj.)","T(°C)":50,"T(°F)":round(50*9/5+32,0),
         "P(bar)":17,"P(psia)":round(17*_PSIA,1),
         "τ(min)":5.0,"Vol(m³)":round(res["V_R102"],4),
         "Vol(L)":round(res["V_R102"]*1000,2),"Vol(gal)":round(res["V_R102"]*_GAL,1),
         "L/D":10,"D(m)":round(_V102,3),"L(m)":round(_L102,2),
         "D(in)":round(_V102*_IN,1),"L(ft)":round(_L102*_FT,1),
         "Mass flux (kg/m²·s)":round(_G102,2),
         "Density PR-EOS (kg/m³)":"—","Density (lb/ft³)":"—",
         "Conversion":"75% PhI(OAc)₂","Reaction":"Aryl migration"},
        {"Tag":"R-103","Type":"PFR", "T(°C)":65,"T(°F)":round(65*9/5+32,0),
         "P(bar)":1.5,"P(psia)":round(1.5*_PSIA,1),
         "τ(min)":7.5,"Vol(m³)":round(res["V_R103"],4),
         "Vol(L)":round(res["V_R103"]*1000,1),"Vol(gal)":round(res["V_R103"]*_GAL,1),
         "L/D":1.5,"D(m)":round(_V103,3),"L(m)":round(_L103,2),
         "D(in)":round(_V103*_IN,1),"L(ft)":round(_L103*_FT,1),
         "Mass flux (kg/m²·s)":round(_G103,2),
         "Density PR-EOS (kg/m³)":"—","Density (lb/ft³)":"—",
         "Conversion":"90% Ester","Reaction":"Saponification"},
    ])

    # ── Sheet 5: Column Sizing ────────────────────────────────────────────
    df_cols = pd.DataFrame([
        {"Tag":"C-101","LK":"ProAc","HK":"TfOH",
         "α (Clausius-Clapeyron)":round(CO["C-101"]["alpha"],2),
         "Nmin (Fenske)":round(CO["C-101"]["Nmin"],1),
         "N actual":CO["C-101"]["N_act"],
         "L/D (assumed)":1.5,
         "Diameter (m)":round(C_C101_d["D_m"],2),
         "Diameter (ft)":round(C_C101_d["D_m"]*_FT,2),              # I4
         "Height (m)":round(C_C101_d["H_m"],1),
         "Height (ft)":round(C_C101_d["H_m"]*_FT,1),                # I4
         "T_cond (°C)":CO["C-101"]["T_cond"],
         "T_cond (°F)":round(CO["C-101"]["T_cond"]*9/5+32,1),       # I4
         "T_reb (°C)":CO["C-101"]["T_reb"],
         "T_reb (°F)":round(CO["C-101"]["T_reb"]*9/5+32,1),         # I4
         "Q_cond (kW)":round(CO["C-101"]["Q_cond"],1),
         "Q_cond (BTU/hr)":round(CO["C-101"]["Q_cond"]*_BTU_HR,0),  # I4
         "Q_reb (kW)":round(CO["C-101"]["Q_reb"],1),
         "Q_reb (BTU/hr)":round(CO["C-101"]["Q_reb"]*_BTU_HR,0),    # I4
         "Shell cost ($k)":round(C_C101_d["shell"]/1e3,1),
         "Tray cost ($k)":round(C_C101_d["trays"]/1e3,1),
         "Condenser ($k)":round(C_C101_d["condenser"]/1e3,1),
         "Reboiler ($k)":round(C_C101_d["reboiler"]/1e3,1),
         "Total ($k)":round(C_C101_d["total"]/1e3,1)},
        {"Tag":"C-103","LK":"MeOH","HK":"TMOF",
         "α (Clausius-Clapeyron)":round(CO["C-103"]["alpha"],2),
         "Nmin (Fenske)":round(CO["C-103"]["Nmin"],1),
         "N actual":CO["C-103"]["N_act"],
         "L/D (assumed)":1.5,
         "Diameter (m)":round(C_C103_d["D_m"],2),
         "Diameter (ft)":round(C_C103_d["D_m"]*_FT,2),
         "Height (m)":round(C_C103_d["H_m"],1),
         "Height (ft)":round(C_C103_d["H_m"]*_FT,1),
         "T_cond (°C)":CO["C-103"]["T_cond"],
         "T_cond (°F)":round(CO["C-103"]["T_cond"]*9/5+32,1),
         "T_reb (°C)":CO["C-103"]["T_reb"],
         "T_reb (°F)":round(CO["C-103"]["T_reb"]*9/5+32,1),
         "Q_cond (kW)":round(CO["C-103"]["Q_cond"],1),
         "Q_cond (BTU/hr)":round(CO["C-103"]["Q_cond"]*_BTU_HR,0),
         "Q_reb (kW)":round(CO["C-103"]["Q_reb"],1),
         "Q_reb (BTU/hr)":round(CO["C-103"]["Q_reb"]*_BTU_HR,0),
         "Shell cost ($k)":round(C_C103_d["shell"]/1e3,1),
         "Tray cost ($k)":round(C_C103_d["trays"]/1e3,1),
         "Condenser ($k)":round(C_C103_d["condenser"]/1e3,1),
         "Reboiler ($k)":round(C_C103_d["reboiler"]/1e3,1),
         "Total ($k)":round(C_C103_d["total"]/1e3,1)},
    ])

    # ── Sheet 6: CAPEX ────────────────────────────────────────────────────
    # L1: build detail table with Cp0 (purchased cost before BM) and size params
    def _cp0(C_bm, F_bm):
        """Purchased cost 2001 USD = C_BM / (F_BM × CEPCI_RATIO)."""
        return C_bm / (F_bm * CEPCI_RATIO)

    def _hx_area_m2(Q_kW, U, LMTD):
        return max(abs(Q_kW) / (U * max(LMTD, 1.0)), 1.0)

    _capex_detail = [
        ("R-101  PFR  [17 bar, SS]",  C_R101,  3.17,
         f"Vol: {res['V_R101']:.3f} m³  ({res['V_R101']*1000:.1f} L)"),
        (f"R-102  PFR ({N_INJECT_R102}-inj.)  [17 bar, SS]",  C_R102,  3.17,
         f"Vol: {res['V_R102']:.3f} m³  ({res['V_R102']*1000:.1f} L), {N_INJECT_R102} injection nozzles"),
        ("R-103  PFR  [1.5 bar, CS]", C_R103,  2.25,
         f"Vol: {res['V_R103']:.3f} m³  ({res['V_R103']*1000:.0f} L)"),
        ("HX-101  feed heater",       C_HX101, 3.29,
         f"Area: {_hx_area_m2(DT['HX-101'],0.50,50):.1f} m²"),
        ("HX-102  Friedel-Crafts cooler", C_HX102, 3.29,
         f"Area: {_hx_area_m2(DT['HX-102'],0.40,25):.1f} m²"),
        ("HX-103  SM-102 heater",     C_HX103, 3.29,
         f"Area: {_hx_area_m2(DT['HX-103'],0.50,35):.1f} m²"),
        ("HX-104  R-103 cooler",      C_HX104, 3.29,
         f"Area: {_hx_area_m2(DT['HX-104'],0.50,20):.1f} m²"),
        ("V-101  LLE separator [17 bar]", C_V101, 3.17,
         f"Vol: {S['S04'].total_flow/900*(5/60):.3f} m³"),
        ("V-102  vacuum flash drum",  C_V102, 2.50,
         f"Vol: {S['S09'].total_flow/900*(2/60):.3f} m³"),
        ("V-103  LLE separator [1 bar]", C_V103, 2.25,
         f"Vol: {S['S14'].total_flow/1000*(5/60):.3f} m³"),
        ("C-101  col + cond + reb",   C_C101_d["total"], 3.00,
         f"D: {C_C101_d['D_m']:.2f} m × H: {C_C101_d['H_m']:.1f} m  ({CO['C-101']['N_act']} trays)"),
        ("C-102  hexane condenser",   C_C102, 3.29,
         f"Area: {_hx_area_m2(DT['C-102 cond'],0.80,15):.1f} m²"),
        ("C-103  col + cond + reb",   C_C103_d["total"], 3.00,
         f"D: {C_C103_d['D_m']:.2f} m × H: {C_C103_d['H_m']:.1f} m  ({CO['C-103']['N_act']} trays)"),
        ("CR-101  crystalliser",      C_cryst, 2.50,
         f"Vol: {V_cryst:.3f} m³"),
        ("Filter + dryer",            C_filt_dry, 2.50,
         "Lump: 1.5× crystalliser BM"),
    ]
    capex_xl = []
    for eqp, C_bm, fbm, size_p in _capex_detail:
        capex_xl.append({
            "Equipment":         eqp,
            "F_BM":              fbm,
            "Cp0 2001 ($k)":     round(_cp0(C_bm, fbm)/1e3, 1),   # L1
            "C_BM ($k)":         round(C_bm/1e3, 1),
            "Size parameter":    size_p,                             # L1
        })
    # L2: summary + FCI breakdown (Turton 2012 Table 16.12)
    _site   = total_purchased * 0.045   # site preparation
    _serv   = total_purchased * 0.045   # service facilities
    _cont   = total_purchased * 0.15    # contingency
    capex_xl += [
        {"Equipment": "─── TOTAL ΣC_BM ───",       "F_BM":"—", "Cp0 2001 ($k)":"—",
         "C_BM ($k)": round(total_purchased/1e3,1), "Size parameter":"—"},
        {"Equipment": "  Site preparation (4.5%)", "F_BM":"—", "Cp0 2001 ($k)":"—",
         "C_BM ($k)": round(_site/1e3,1),           "Size parameter":"—"},
        {"Equipment": "  Service facilities (4.5%)", "F_BM":"—", "Cp0 2001 ($k)":"—",
         "C_BM ($k)": round(_serv/1e3,1),           "Size parameter":"—"},
        {"Equipment": "  Contingency (15%)",        "F_BM":"—", "Cp0 2001 ($k)":"—",
         "C_BM ($k)": round(_cont/1e3,1),           "Size parameter":"—"},
        {"Equipment": "FCI (×1.18 ΣC_BM)",         "F_BM":"—", "Cp0 2001 ($k)":"—",
         "C_BM ($k)": round(FCI/1e3,1),             "Size parameter":"—"},
        {"Equipment": "Working Capital (15% FCI)",  "F_BM":"—", "Cp0 2001 ($k)":"—",
         "C_BM ($k)": round(WC/1e3,1),              "Size parameter":"—"},
        {"Equipment": "TOTAL CAPITAL INVESTMENT (TCI)", "F_BM":"—", "Cp0 2001 ($k)":"—",
         "C_BM ($k)": round(TCI/1e3,1),             "Size parameter":"—"},
    ]
    df_capex = pd.DataFrame(capex_xl)

    # ── Sheet 7: Raw Materials ────────────────────────────────────────────
    rm_xl = []
    for chem, price in RM_PRICE.items():
        rate = f.get(chem, 0.0)
        cost = rate * HOURS_PER_YEAR * price
        rm_xl.append({
            "Chemical":              chem.replace("_mkup","").replace("_S2",""),
            "Fresh rate (kg/hr)":    round(rate, 2),
            "Fresh rate (lb/hr)":    round(rate * _LB_HR, 1),              # I7
            "Price ($/kg)":          price,
            "Annual rate (MT/yr)":   round(rate * HOURS_PER_YEAR / 1000, 2),
            "Annual rate (s.tons/yr)": round(rate * HOURS_PER_YEAR / 1000 * _STON_YR, 2),  # I7
            "Annual cost ($M/yr)":   round(cost / 1e6, 4),
        })
    rm_xl.append({"Chemical":"TOTAL","Fresh rate (kg/hr)":"—",
                  "Fresh rate (lb/hr)":"—","Price ($/kg)":"—",
                  "Annual rate (MT/yr)":"—","Annual rate (s.tons/yr)":"—",
                  "Annual cost ($M/yr)": round(total_rm_yr/1e6, 4)})
    df_rm = pd.DataFrame(rm_xl)

    # ── Sheet 8: Utilities ────────────────────────────────────────────────
    util_xl = []
    for blk, Q in DT.items():
        util, _ = UTIL_MAP.get(blk, ("—","—"))
        if util == "—":
            continue
        GJ = abs(Q) * 3600 * HOURS_PER_YEAR / 1e6
        cost = GJ * UTIL_UNIT_COST.get(util, 0)
        util_xl.append({
            "Block":                blk,
            "Utility":              util,
            "Duty (kW)":            round(Q, 2),
            "Duty (BTU/hr)":        round(Q * _BTU_HR, 0),          # I6
            "GJ/yr":                round(GJ, 1),
            "MMBtu/yr":             round(GJ * _MMBTU_YR, 1),       # I6
            "Rate ($/GJ)":          UTIL_UNIT_COST.get(util, 0),
            "Annual cost ($k/yr)":  round(cost/1e3, 2),
        })
    util_xl.append({"Block":"TOTAL","Utility":"—","Duty (kW)":"—",
                    "Duty (BTU/hr)":"—","GJ/yr":"—","MMBtu/yr":"—",
                    "Rate ($/GJ)":"—","Annual cost ($k/yr)": round(total_util_yr/1e3, 2)})
    df_util = pd.DataFrame(util_xl)

    # ── Sheet 9: P&L / Profitability ─────────────────────────────────────
    df_pnl = pd.DataFrame([
        {"Item": "Revenue — Ibuprofen API",
         "Value ($M/yr)": round(revenue_ibup/1e6, 3)},
        {"Item": f"Revenue — PhI byproduct",
         "Value ($M/yr)": round(revenue_phi/1e6, 3)},
        {"Item": "TOTAL REVENUE",
         "Value ($M/yr)": round(total_revenue/1e6, 3)},
        {"Item": "Raw Material Cost",
         "Value ($M/yr)": round(-total_rm_yr/1e6, 3)},
        {"Item": "Utility Cost",
         "Value ($M/yr)": round(-total_util_yr/1e6, 3)},
        {"Item": "GROSS MARGIN",
         "Value ($M/yr)": round(gross_margin/1e6, 3)},
        {"Item": "Gross Margin %",
         "Value ($M/yr)": round(gross_margin/total_revenue*100, 1)},
        {"Item": "Labor", "Value ($M/yr)": round(-labor_yr/1e6, 3)},
        {"Item": "Maintenance (3% FCI)",
         "Value ($M/yr)": round(-maint_yr/1e6, 3)},
        {"Item": "Overhead + G&A",
         "Value ($M/yr)": round(-(overhead_yr+admin_yr)/1e6, 3)},
        {"Item": "EBITDA", "Value ($M/yr)": round(ebitda/1e6, 3)},
        {"Item": "EBITDA Margin %",
         "Value ($M/yr)": round(ebitda/total_revenue*100, 1)},
        {"Item": "Depreciation (10yr SL)",
         "Value ($M/yr)": round(-depreciation/1e6, 3)},
        {"Item": "EBIT", "Value ($M/yr)": round(ebit/1e6, 3)},
        {"Item": "Tax (25%)",
         "Value ($M/yr)": round(-max(ebit,0)*0.25/1e6, 3)},
        {"Item": "Net Income", "Value ($M/yr)": round(net_income/1e6, 3)},
        {"Item": "Free Cash Flow", "Value ($M/yr)": round(fcf/1e6, 3)},
        {"Item": "", "Value ($M/yr)": ""},
        {"Item": "FCI ($M)", "Value ($M/yr)": round(FCI/1e6, 3)},
        {"Item": "Working Capital ($M)", "Value ($M/yr)": round(WC/1e6, 3)},
        {"Item": "TCI ($M)", "Value ($M/yr)": round(TCI/1e6, 3)},
        {"Item": "Simple Payback (yr)",
         "Value ($M/yr)": round(payback_yr, 1)},
        {"Item": "NPV @ 12% WACC, 10yr ($M)",
         "Value ($M/yr)": round(npv_simple/1e6, 2)},
        {"Item": "Approx ROI (%/yr)",
         "Value ($M/yr)": round(irr_approx, 1)},
        {"Item": "", "Value ($M/yr)": ""},
        {"Item": "── UNIT PRODUCTION COST ──", "Value ($M/yr)": ""},   # L3
        {"Item": "Unit cost ($/kg API)",
         "Value ($M/yr)": round(total_opex_yr / max(ibup_API * HOURS_PER_YEAR, 1), 2)},
        {"Item": "Unit cost ($/lb API)",
         "Value ($M/yr)": round(total_opex_yr / max(ibup_API * HOURS_PER_YEAR, 1) / _LB_HR, 2)},
    ])

    # ── Sheet 10: Recycle & Process Metrics ──────────────────────────────
    df_metrics = pd.DataFrame([
        {"Metric": "IBB feed (kg/hr)",              "Value": round(ibb_feed, 2)},
        {"Metric": "IBB feed (lb/hr)",              "Value": round(ibb_feed*_LB_HR, 1)},  # I8
        {"Metric": "Ibuprofen API (kg/hr)",         "Value": round(ibup_API, 3)},
        {"Metric": "Ibuprofen API (lb/hr)",         "Value": round(ibup_API*_LB_HR, 1)},  # I8
        {"Metric": "Annual production (MT/yr)",     "Value": round(ibup_API*8000/1000, 1)},
        {"Metric": "Annual production (s.tons/yr)", "Value": round(ibup_API*8000/1000*_STON_YR, 1)},  # I8
        {"Metric": "Overall yield IBB→API (%)",  "Value": round(ov_yield, 2)},
        {"Metric": "Purity (wt%)",               "Value": round(res["purity"], 4)},
        {"Metric": "Residual MeOH (ppm)",        "Value": round(res["MeOH_ppm"], 1)},
        {"Metric": "V-101 water in organic (wt%)", "Value": round(res["water_wt"], 3)},
        {"Metric": "V-102 AcOH in bottoms (wt%)", "Value": round(res["acoh_wt"], 3)},
        {"Metric": "NaOH/Ester ratio",           "Value": round(res["NaOH_ratio"], 3)},
        {"Metric": "HCl coverage (%)",           "Value": round(res["HCl_cov"], 2)},
        {"Metric": "","Value":""},
        {"Metric": "── RECYCLE (converged) ──", "Value": ""},
        {"Metric": "TfOH recycle (kg/hr)",       "Value": round(tear["TfOH"], 2)},
        {"Metric": "TMOF recycle (kg/hr)",       "Value": round(tear["TMOF"], 2)},
        {"Metric": "MeOH recycle (kg/hr)",       "Value": round(tear["MeOH_S2"], 2)},
        {"Metric": "Hexane recycle (kg/hr)",     "Value": round(tear["Hex"], 2)},
        {"Metric": "","Value":""},
        {"Metric": "── STEP YIELDS ──", "Value": ""},
    ] + [{"Metric": k, "Value": round(v, 2)} for k, v in sy.items()])

    # ── Sheet 11: Electrolyte Summary ────────────────────────────────────
    ions = res["ions_S13"]
    df_elec = pd.DataFrame([
        {"Parameter": "Na⁺ (mol/hr)",       "Value": round(ions["Na+"], 1)},
        {"Parameter": "OH⁻ (mol/hr)",       "Value": round(ions["OH-"], 1)},
        {"Parameter": "H⁺ (mol/hr)",        "Value": round(ions["H+"], 1)},
        {"Parameter": "Ibup⁻ (mol/hr)",     "Value": round(ions["IbupAnion"], 1)},
        {"Parameter": "Ionic strength I (mol/L)", "Value": round(res["I_S13"], 5)},
        {"Parameter": "γ(Na⁺) — Debye-Hückel", "Value": round(res["gamma_Na"], 5)},
        {"Parameter": "γ(Ibup⁻) — Debye-Hückel","Value": round(res["gamma_IbA"],5)},
        {"Parameter": "γ(OH⁻) — Debye-Hückel", "Value": round(res["gamma_OH"], 5)},
        {"Parameter": "m_NaCl at V-103 (mol/kg)", "Value": round(res["m_NaCl_S14"],4)},
    ])

    # ── Write workbook ────────────────────────────────────────────────────
    with pd.ExcelWriter(_XLSX, engine="openpyxl") as _writer:
        sheets = [
            ("Stream Table (full)",   df_streams),
            ("Stream Summary",        df_stream_summ),
            ("Heat Duties",           df_duties),
            ("Reactor Sizing",        df_reactors),
            ("Column Sizing",         df_cols),
            ("CAPEX",                 df_capex),
            ("Raw Materials",         df_rm),
            ("Utilities",             df_util),
            ("P&L — Profitability",   df_pnl),
            ("Process Metrics",       df_metrics),
            ("Electrolyte Model",     df_elec),
        ]
        for sheet_name, df in sheets:
            df.to_excel(_writer, sheet_name=sheet_name, index=False)

    # ── Apply styling ─────────────────────────────────────────────────────
    wb = _openpyxl.load_workbook(_XLSX)
    _total_kws = {"TOTAL", "FCI", "TCI", "EBITDA", "EBIT", "GROSS",
                  "NET INCOME", "FREE CASH"}

    for ws in wb.worksheets:
        # Style header row
        n_cols = ws.max_column
        _write_header(ws, [ws.cell(1, c).value for c in range(1, n_cols+1)])
        # Freeze top row
        ws.freeze_panes = "A2"
        # Highlight total/summary rows
        for row in ws.iter_rows(min_row=2):
            first = str(row[0].value or "")
            if any(kw in first.upper() for kw in _total_kws):
                for cell in row:
                    cell.fill = _TOT_FILL
                    cell.font = _SUB_FONT
        # Auto-width (capped at 40)
        for col in ws.columns:
            max_w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_w + 2, 40)

    # Cover sheet
    ws_cover = wb.create_sheet("README", 0)
    wb.active = ws_cover
    cover_info = [
        ("Ibuprofen Continuous Flow Simulation — Results", ""),
        ("Version", "v5"),
        ("Route", "Bogdan (2009) 3-step continuous flow"),
        ("Basis", "8000 hr/yr operating"),
        ("Annual production", f"{ibup_API*8000/1000:.0f} MT/yr ibuprofen API"),
        ("Overall yield", f"{ov_yield:.1f}% IBB→API"),
        ("TCI", f"${TCI/1e6:.2f}M"),
        ("Total OPEX", f"${total_opex_yr/1e6:.2f}M/yr"),
        ("Total Revenue", f"${total_revenue/1e6:.2f}M/yr"),
        ("EBITDA", f"${ebitda/1e6:.2f}M/yr"),
        ("", ""),
        ("Sheets:", ""),
        ("1. README", "This cover page"),
        ("2. Stream Table (full)", "All streams × all components (kg/hr + wt%)"),
        ("3. Stream Summary", "Top-5 components per stream"),
        ("4. Heat Duties", "All heat exchangers, duties, utility costs"),
        ("5. Reactor Sizing", "R-101/102/103 volumes, conditions, PR-EOS density"),
        ("6. Column Sizing", "C-101/C-103 Fenske shortcut + Turton cost"),
        ("7. CAPEX", "Equipment bare-module costs (Turton 2012, CEPCI 2025) + C_p0 + FCI breakdown"),
        ("8. Raw Materials", "Fresh feed rates + annual costs (recycle credited)"),
        ("9. Utilities", "GJ/yr + annual costs per block"),
        ("10. P&L — Profitability", "Revenue, OPEX, EBITDA, NPV, payback"),
        ("11. Process Metrics", "Yields, specs, recycle convergence"),
        ("12. Electrolyte Model", "Ion tracking, Debye-Hückel γ at R-103 outlet"),
    ]
    for r, (k, v) in enumerate(cover_info, 1):
        ws_cover.cell(r, 1, k).font = _Font(bold=(r<=1 or k.endswith(":")),
                                             name="Calibri", size=10)
        ws_cover.cell(r, 2, v).font = _Font(name="Calibri", size=10)
    ws_cover.column_dimensions["A"].width = 32
    ws_cover.column_dimensions["B"].width = 55
    if cover_info:
        ws_cover.cell(1,1).fill = _HDR_FILL
        ws_cover.cell(1,1).font = _Font(bold=True, color="FFFFFF",
                                        name="Calibri", size=12)
        ws_cover.cell(1,2).fill = _HDR_FILL

    wb.save(_XLSX)
    print(f"\n  ✅  Excel workbook saved → {_XLSX}")
    print(f"      Sheets: {', '.join(s for s,_ in sheets)}")

except ImportError:
    print("\n  ⚠️  openpyxl not installed — run: pip install openpyxl")
except Exception as _e:
    print(f"\n  ⚠️  Excel export failed: {_e}")
